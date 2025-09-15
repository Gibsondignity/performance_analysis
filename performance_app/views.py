from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect
from datetime import datetime, timedelta
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import json
import csv
from django.db.models import Count
from .models import *
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models.functions import TruncMonth
from collections import defaultdict
from django.utils.timezone import now
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .forms import UserForm, EmployeeProfileForm, EditUserForm, PerformanceRecordForm, EvaluationForm
from .models import Evaluation, Attendance, Task, PeerReview, Training, EvaluationCriteria
from ai_engine.inference_service import AIService
from django.contrib import messages
from django.db import IntegrityError
from django.db.models import Avg, ExpressionWrapper, F, FloatField, Sum
from functools import wraps
from datetime import datetime
from django.utils.timezone import localtime
import csv
from django.db.models import Q


# ROLE-BASED ACCESS CONTROL DECORATORS
def role_required(allowed_roles):
    """
    Decorator to restrict access based on user roles
    """
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')

            if request.user.role not in allowed_roles:
                messages.error(request, f'❌ Access denied. Required roles: {", ".join(allowed_roles)}')
                return redirect('dashboard')

            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator


def hr_or_admin_required(view_func):
    """Decorator for HR and admin access"""
    return role_required(['HR', 'HIGH_MANAGER'])(view_func)


def manager_or_above_required(view_func):
    """Decorator for manager and above access"""
    return role_required(['MANAGER', 'MIDDLE_MANAGER', 'HIGH_MANAGER', 'HR'])(view_func)


def employee_access_only(view_func):
    """Decorator to ensure only employees can access their own data"""
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')

        # Allow HR and managers to access, but restrict employees to their own data
        if request.user.role == 'EMPLOYEE':
            # This will be checked in the view itself
            pass

        return view_func(request, *args, **kwargs)
    return _wrapped_view


def custom_login_view(request):
    error = None
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        password = request.POST.get('password')
        user = authenticate(request, username=employee_id, password=password)

        if user:
            login(request, user)
            profile = EmployeeProfile.objects.filter(user=user).first()
            # Fetch employee profile details
            if profile:
                print(f"Profile found: {profile}")
                request.session['employee_id'] = user.employee_id
                request.session['first_name'] = profile.first_name or ''
                request.session['last_name'] = profile.last_name or ''
            else:
                # Profile doesn't exist yet
                request.session['employee_id'] = user.employee_id
                request.session['first_name'] = ''
                request.session['last_name'] = ''

            if user.role == 'EMPLOYEE':
                return redirect('employee')
            else:
                return redirect('dashboard')

        error = "Invalid Employee ID or password"

    return render(request, 'login.html', {'error': error, 'year': now().year})



def logout_view(request):
    if request.method == 'POST':
        logout(request)
        request.session.flush()
        return redirect('login')
    return render(request, 'login.html', {'year': datetime.now().year})


@login_required
def analytical_dashboard(request):
    """
    Manager dashboard with comprehensive team analytics
    """
    user = request.user

    # Get team members based on role
    if user.role == 'MANAGER':
        team_members = EmployeeProfile.objects.filter(user__role='EMPLOYEE')
    elif user.role == 'MIDDLE_MANAGER':
        team_members = EmployeeProfile.objects.filter(user__role__in=['EMPLOYEE', 'MANAGER'])
    elif user.role in ['HIGH_MANAGER', 'HR']:
        team_members = EmployeeProfile.objects.all()
    else:
        team_members = EmployeeProfile.objects.none()

    # Team Performance Overview
    recent_records = PerformanceRecord.objects.filter(
        employee__in=team_members
    ).select_related('employee__user').order_by('-performance_end_date')[:50]

    # Calculate average scores per employee using new KPI system
    employee_performance = {}
    for record in recent_records:
        emp_id = record.employee.id
        if emp_id not in employee_performance:
            employee_performance[emp_id] = {
                'name': f"{record.employee.first_name} {record.employee.last_name}",
                'scores': [],
                'average_score': 0
            }
        # Use new KPI scoring method if available
        if hasattr(record, 'get_overall_kpi_score'):
            overall_score = record.get_overall_kpi_score()
        else:
            # Fallback to legacy calculation
            overall_score = record.sales_achieved_percent
        employee_performance[emp_id]['scores'].append(overall_score)

    # Calculate averages
    for emp_data in employee_performance.values():
        if emp_data['scores']:
            emp_data['average_score'] = sum(emp_data['scores']) / len(emp_data['scores'])

    # Prepare team performance data for chart
    team_performance_data = {
        'labels': [],
        'scores': []
    }

    for emp_data in employee_performance.values():
        if emp_data['scores']:
            avg_score = sum(emp_data['scores']) / len(emp_data['scores'])
            team_performance_data['labels'].append(emp_data['name'])
            team_performance_data['scores'].append(round(avg_score, 1))

    # Top and Low Performers
    def calculate_average(scores):
        return sum(scores) / len(scores) if scores else 0

    top_performers = sorted(employee_performance.items(),
                          key=lambda x: calculate_average(x[1]['scores']),
                          reverse=True)[:5]

    low_performers = sorted(employee_performance.items(),
                            key=lambda x: calculate_average(x[1]['scores']))[:5]

    # Add average to each employee data
    for emp_id, emp_data in employee_performance.items():
        emp_data['average_score'] = calculate_average(emp_data['scores'])

    # Feedback Sentiment Analysis
    evaluations = Evaluation.objects.filter(employee__in=team_members).order_by('-date')[:100]

    sentiment_counts = {'positive': 0, 'neutral': 0, 'negative': 0}
    for eval in evaluations:
        if eval.remarks:
            text = eval.remarks.lower()
            if any(word in text for word in ['excellent', 'great', 'good', 'improved', 'strong']):
                sentiment_counts['positive'] += 1
            elif any(word in text for word in ['poor', 'weak', 'needs improvement', 'concerning']):
                sentiment_counts['negative'] += 1
            else:
                sentiment_counts['neutral'] += 1

    # Attendance Heatmap Data
    from django.utils import timezone
    current_month = timezone.now().replace(day=1)
    attendance_records = Attendance.objects.filter(
        employee__in=team_members,
        date__gte=current_month
    ).select_related('employee')

    attendance_heatmap = {}
    for record in attendance_records:
        emp_name = f"{record.employee.first_name} {record.employee.last_name}"
        if emp_name not in attendance_heatmap:
            attendance_heatmap[emp_name] = {'present': 0, 'absent': 0, 'late': 0}

        if record.is_present:
            attendance_heatmap[emp_name]['present'] += 1
        else:
            attendance_heatmap[emp_name]['absent'] += 1
        if record.is_late:
            attendance_heatmap[emp_name]['late'] += 1

    # Calculate attendance rates
    for emp_name, data in attendance_heatmap.items():
        total = data['present'] + data['absent']
        if total > 0:
            data['attendance_rate'] = round((data['present'] / total) * 100, 1)
        else:
            data['attendance_rate'] = 0
        data['total'] = total

    # AI Insights Summary
    total_evaluations = evaluations.count()
    avg_team_score = sum(eval.performance_score for eval in evaluations) / total_evaluations if total_evaluations > 0 else 0

    # Calculate real attendance metrics
    current_month = timezone.now().replace(day=1)
    monthly_attendance = Attendance.objects.filter(
        employee__in=team_members,
        date__gte=current_month
    )

    attendance_rate = 0
    if monthly_attendance.exists():
        total_attendance_records = monthly_attendance.count()
        present_records = monthly_attendance.filter(is_present=True).count()
        attendance_rate = (present_records / total_attendance_records * 100) if total_attendance_records > 0 else 0

    # Calculate active trainings
    active_trainings = Training.objects.filter(
        employee__in=team_members,
        status__in=['PLANNED', 'IN_PROGRESS']
    ).count()

    ai_insights = {
        'total_evaluations': total_evaluations,
        'avg_team_score': round(avg_team_score, 1),
        'sentiment_distribution': sentiment_counts,
        'top_performer': top_performers[0][1]['name'] if top_performers else 'N/A',
        'needs_attention': len([p for p in employee_performance.values() if p['average_score'] < 60]),
        'attendance_rate': round(attendance_rate, 1),
        'active_trainings': active_trainings
    }

    # Calculate team trend data for the last 12 months
    from django.db.models.functions import ExtractMonth, ExtractYear

    # Get the raw data and calculate performance in Python
    trend_records = PerformanceRecord.objects.filter(
        employee__in=team_members,
        performance_end_date__gte=timezone.now() - timezone.timedelta(days=365)
    ).annotate(
        month=ExtractMonth('performance_end_date'),
        year=ExtractYear('performance_end_date')
    ).values('year', 'month', 'sales_volume', 'sales_target')

    # Group by year and month, calculate average performance
    trend_data = {}
    for record in trend_records:
        key = f"{record['year']}-{record['month']:02d}"
        if key not in trend_data:
            trend_data[key] = []

        # Calculate performance percentage
        if record['sales_target'] and record['sales_target'] > 0:
            performance = (record['sales_volume'] / record['sales_target']) * 100
        else:
            performance = 0
        trend_data[key].append(performance)

    # Calculate averages
    team_trend_data = []
    for key, performances in trend_data.items():
        if performances:
            avg_performance = sum(performances) / len(performances)
            year, month = key.split('-')
            team_trend_data.append({
                'year': int(year),
                'month': int(month),
                'avg_performance': avg_performance
            })

    # Sort by year and month
    team_trend_data.sort(key=lambda x: (x['year'], x['month']))

    # Create trend data
    trend_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    trend_scores = []

    current_year = timezone.now().year
    for i, month in enumerate(trend_labels, 1):
        # Find data for this month
        month_data = None
        for data in team_trend_data:
            if data['year'] == current_year and data['month'] == i:
                month_data = data
                break

        if month_data and month_data['avg_performance']:
            trend_scores.append(round(month_data['avg_performance'], 1))
        else:
            # Use current average or 0
            trend_scores.append(round(ai_insights['avg_team_score'], 1) if ai_insights['avg_team_score'] > 0 else 0)

    context = {
        'team_performance_data': team_performance_data,
        'top_performers': top_performers,
        'low_performers': low_performers,
        'sentiment_data': sentiment_counts,
        'attendance_heatmap': attendance_heatmap,
        'ai_insights': ai_insights,
        'team_size': team_members.count(),
        'total_evaluations': total_evaluations,
        'trend_labels': trend_labels,
        'trend_scores': trend_scores
    }

    return render(request, 'dashboard/manager/manager_dashboard.html', context)





# HR MANAGER DASHBOARD
@hr_or_admin_required
def employee_management(request):

    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        if user_id:
            user = get_object_or_404(User, pk=user_id)
            user_form = UserForm(request.POST, instance=user)
            profile_form = EmployeeProfileForm(request.POST, instance=user.employeeprofile)
        else:
            user_form = UserForm(request.POST)
            profile_form = EmployeeProfileForm(request.POST)
        print(user_form.errors, profile_form.errors)
        # contact_form = ContactInfoForm(request.POST)
        if user_form.is_valid() and profile_form.is_valid():
            user = user_form.save()
            profile = profile_form.save(commit=False)
            profile.user = user
            profile.save()
            messages.success(request, 'Employee saved successfully.')
            return redirect('employee_management')
        else:
            messages.error(request, 'There was an error saving the employee.')

    user_form = UserForm()
    profile_form = EmployeeProfileForm()
    # contact_form = ContactInfoForm()

    employees = EmployeeProfile.objects.select_related('user')
    return render(request, 'dashboard/hr/employee_mgmt.html', {
        'employees': employees,
        'user_form': user_form,
        'profile_form': profile_form,
        # 'contact_form': contact_form,
    })



@hr_or_admin_required
def add_employee(request):

    if request.method == 'POST':
        user_form = UserForm(request.POST)
        profile_form = EmployeeProfileForm(request.POST)

        if user_form.is_valid() and profile_form.is_valid():
            try:
                # Create user with auto-generated employee_id using CustomUserManager
                user_data = user_form.cleaned_data
                user = User.objects.create_user(
                    employee_id=None,  # Will be auto-generated
                    password=user_data.get('password'),
                    role=user_data.get('role'),
                    branch=user_data.get('branch'),
                    is_active=user_data.get('is_active', True)
                )

                profile = profile_form.save(commit=False)
                profile.user = user
                profile.created_by = request.user
                profile.save()
                messages.success(request, f'✅ Employee added successfully. Employee ID: {user.employee_id}')
                return redirect('employee_management')
            except IntegrityError as e:
                if 'unique constraint' in str(e).lower() or 'UNIQUE constraint failed' in str(e):
                    messages.error(request, '❌ Employee ID already exists.')
                else:
                    messages.error(request, f'❌ An error occurred: {e}')
        else:
            # Forms not valid; let the template display field errors
            messages.error(request, '❌ Please correct the errors below.')
    else:
        user_form = UserForm()
        profile_form = EmployeeProfileForm()

    return render(request, 'dashboard/hr/add_employee.html', {
        'user_form': user_form,
        'profile_form': profile_form,
    })


@login_required
def view_employee(request, user_id):
    employee = get_object_or_404(EmployeeProfile, user__id=user_id)
    return render(request, 'dashboard/hr/view_employee.html', {'employee': employee})



@login_required
def edit_employee(request, user_id):
    user = get_object_or_404(User, id=user_id)
    profile = get_object_or_404(EmployeeProfile, user=user)

    if request.method == 'POST':
        user_form = EditUserForm(request.POST, instance=user)
        profile_form = EmployeeProfileForm(request.POST, instance=profile)

        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, 'Employee updated successfully.')
            return redirect('employee_management')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        user_form = EditUserForm(instance=user)
        profile_form = EmployeeProfileForm(instance=profile)

    return render(request, 'dashboard/hr/edit_employee.html', {
        'user_form': user_form,
        'profile_form': profile_form,
        'employee': profile
    })



@login_required
def delete_employee(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        user.delete()
        messages.success(request, 'Employee deleted successfully.')
        return redirect('employee_management')

    messages.error(request, 'Invalid request method.')
    return redirect('employee_management')


@login_required
def performance_records(request):
    records = PerformanceRecord.objects.all()
    return render(request, 'dashboard/hr/performance_records.html', {'records': records})







# LOW LEVEL VIEWS FOR PERFORMANCE RECORDS
@login_required
def performance_record_list(request):
    """
    List all performance records with optional filtering by employee and performance date range.
    Low level managers should only see employees, not other managers.
    """
    user = request.user

    # Base queryset for performance records
    records = PerformanceRecord.objects.select_related('employee', 'employee__user').order_by('-performance_start_date')

    # Base queryset for employees
    employees = EmployeeProfile.objects.select_related('user')

    # Restrict employee list for Low Level Managers
    if user.role == 'MANAGER':
        employees = employees.filter(user__role='EMPLOYEE')  # Only show actual employees
    elif user.role == 'MIDDLE_MANAGER':
        employees = employees.filter(user__role='MANAGER')
    elif user.role == 'HIGH_LEVEL_MANAGER':
        employees = employees.filter(user__role='MIDDLE_MANAGER')
    else:
        employees = None
    

    # Get filter values from GET request
    employee_id = request.GET.get('employee')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Apply filters
    if employee_id:
        records = records.filter(employee_id=employee_id)
    if start_date and end_date:
        records = records.filter(
            performance_start_date__gte=start_date,
            performance_end_date__lte=end_date
        )

    # Pass data to template
    context = {
        'records': records,
        'employees': employees,
        'selected_employee': int(employee_id) if employee_id else None,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'dashboard/low_level_manager/performance_records.html', context)



@login_required
def add_performance_record(request):
    """
    Add a new performance record with flexible KPI support.
    """
    if request.method == 'POST':
        form = PerformanceRecordForm(request.POST)
        if form.is_valid():
            performance_record = form.save(commit=False)

            # Handle flexible KPI values
            kpi_values = {}
            role_kpis = KPI.objects.filter(role=performance_record.employee.user.role, is_active=True)

            for kpi in role_kpis:
                target_key = f'kpi_{kpi.id}_target'
                actual_key = f'kpi_{kpi.id}_actual'

                target = request.POST.get(target_key)
                actual = request.POST.get(actual_key)

                if target and actual:
                    try:
                        kpi_values[kpi.name] = {
                            'target': float(target),
                            'actual': float(actual)
                        }
                    except ValueError:
                        pass

            performance_record.kpi_values = kpi_values
            performance_record.created_by = request.user
            performance_record.save()

            messages.success(request, '✅ Performance record added successfully.')
            return redirect('performance_record_list')
        else:
            messages.error(request, '❌ Please correct the errors below.')
    else:
        form = PerformanceRecordForm()

    # Get KPIs for the selected employee (if any)
    employee_id = request.GET.get('employee')
    role_kpis = []
    if employee_id:
        try:
            employee = EmployeeProfile.objects.get(id=employee_id)
            role_kpis = KPI.objects.filter(role=employee.user.role, is_active=True)
        except EmployeeProfile.DoesNotExist:
            messages.warning(request, '⚠️ Selected employee profile not found.')

    return render(request, 'dashboard/low_level_manager/add_performance.html', {
        'form': form,
        'role_kpis': role_kpis
    })


@login_required
def edit_performance_record(request, pk):
    """
    Edit an existing performance record with flexible KPI support.
    """
    record = get_object_or_404(PerformanceRecord, pk=pk)

    if request.method == 'POST':
        form = PerformanceRecordForm(request.POST, instance=record)
        if form.is_valid():
            performance_record = form.save(commit=False)

            # Handle flexible KPI values
            kpi_values = {}
            role_kpis = KPI.objects.filter(role=performance_record.employee.user.role, is_active=True)

            for kpi in role_kpis:
                target_key = f'kpi_{kpi.id}_target'
                actual_key = f'kpi_{kpi.id}_actual'

                target = request.POST.get(target_key)
                actual = request.POST.get(actual_key)

                if target and actual:
                    try:
                        kpi_values[kpi.name] = {
                            'target': float(target),
                            'actual': float(actual)
                        }
                    except ValueError:
                        pass

            performance_record.kpi_values = kpi_values
            performance_record.updated_by = request.user
            performance_record.save()

            messages.success(request, '✅ Performance record updated successfully.')
            return redirect('performance_record_list')
        else:
            messages.error(request, '❌ Please correct the errors below.')
    else:
        form = PerformanceRecordForm(instance=record)

    # Get KPIs for the employee's role
    role_kpis = KPI.objects.filter(role=record.employee.user.role, is_active=True)

    return render(request, 'dashboard/low_level_manager/edit_performance_records.html', {
        'form': form,
        'record': record,
        'role_kpis': role_kpis
    })


@login_required
def delete_performance_record(request, pk):
    """
    Delete a performance record.
    """
    record = get_object_or_404(PerformanceRecord, pk=pk)

    if request.method == 'POST':
        record.delete()
        messages.success(request, '🗑️ Performance record deleted successfully.')
        return redirect('performance_record_list')
    
    messages.error(request, '❌ Invalid request method.')

    return redirect('performance_record_list')







def analytics_view(request):
    employees = EmployeeProfile.objects.select_related('user')
    selected_employee = request.GET.get('employee')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    records = PerformanceRecord.objects.select_related('employee').all()

    if selected_employee:
        records = records.filter(employee_id=selected_employee)
    if start_date:
        records = records.filter(data_created__gte=start_date)
    if end_date:
        records = records.filter(data_created__lte=end_date)

    # Manually compute achieved percentages since they're not DB fields
    def safe_percent(numerator, denominator):
        return (numerator / denominator * 100) if denominator else 0

    sales_percentages = [safe_percent(r.sales_volume, r.sales_target) for r in records]
    dist_percentages = [safe_percent(r.distribution_volume, r.distribution_target) for r in records]
    revenue_percentages = [safe_percent(r.revenue_volume, r.revenue_target) for r in records]
    engagement_scores = [r.team_engagement_score for r in records]

    avg_sales = round(sum(sales_percentages) / len(sales_percentages), 2) if sales_percentages else 0
    avg_distribution = round(sum(dist_percentages) / len(dist_percentages), 2) if dist_percentages else 0
    avg_revenue = round(sum(revenue_percentages) / len(revenue_percentages), 2) if revenue_percentages else 0
    avg_engagement = round(sum(engagement_scores) / len(engagement_scores), 1) if engagement_scores else 0

    # For chart display
    chart_labels = [r.data_created.strftime('%Y-%m-%d') for r in records]
    chart_sales = sales_percentages
    chart_revenue = revenue_percentages
    chart_engagement = engagement_scores

    # Calculate dynamic status messages
    def get_performance_status(value, target=100):
        """Calculate performance status based on value"""
        if value >= target * 0.95:  # 95% or above
            return "Above target"
        elif value >= target * 0.85:  # 85% or above
            return "On target"
        elif value >= target * 0.70:  # 70% or above
            return "Below target"
        else:
            return "Needs improvement"

    def get_engagement_status(score):
        """Calculate engagement status based on score out of 10"""
        if score >= 8:
            return "Excellent engagement"
        elif score >= 6:
            return "Good engagement"
        elif score >= 4:
            return "Fair engagement"
        else:
            return "Needs improvement"

    # Calculate dynamic AI insights
    def get_team_performance_insight(avg_sales, avg_distribution, avg_revenue, avg_engagement):
        """Generate dynamic team performance insight"""
        overall_avg = (avg_sales + avg_distribution + avg_revenue) / 3
        if overall_avg >= 90:
            return "Your team is performing exceptionally well across all key metrics."
        elif overall_avg >= 80:
            return "Your team is performing above average across all key metrics."
        elif overall_avg >= 70:
            return "Your team is performing adequately with room for improvement."
        else:
            return "Your team needs focused attention to improve performance metrics."

    def get_growth_trend_insight(chart_sales, chart_revenue):
        """Generate dynamic growth trend insight"""
        if len(chart_sales) >= 2 and len(chart_revenue) >= 2:
            sales_trend = chart_sales[-1] - chart_sales[0] if chart_sales[0] > 0 else 0
            revenue_trend = chart_revenue[-1] - chart_revenue[0] if chart_revenue[0] > 0 else 0

            if sales_trend > 10 and revenue_trend > 10:
                return "Strong upward trend in both sales and revenue over the selected period."
            elif sales_trend > 5 or revenue_trend > 5:
                return "Consistent improvement in sales and revenue over the selected period."
            elif sales_trend < -5 or revenue_trend < -5:
                return "Declining trend in sales and revenue requires immediate attention."
            else:
                return "Stable performance with minor fluctuations in sales and revenue."
        return "Consistent improvement in sales and revenue over the selected period."

    def get_recommendation_insight(avg_engagement, avg_sales, avg_distribution, avg_revenue):
        """Generate dynamic recommendation insight"""
        if avg_engagement < 6:
            return "Consider team-building activities to further boost engagement scores."
        elif avg_sales < 80:
            return "Focus on sales training and target setting to improve achievement rates."
        elif avg_distribution < 80:
            return "Review distribution processes and provide additional resources if needed."
        elif avg_revenue < 80:
            return "Implement revenue optimization strategies and monitor key drivers."
        else:
            return "Continue current strategies while monitoring for further optimization opportunities."

    context = {
        'employees': employees,
        'selected_employee': int(selected_employee) if selected_employee else '',
        'start_date': start_date,
        'end_date': end_date,
        'avg_sales': avg_sales,
        'avg_distribution': avg_distribution,
        'avg_revenue': avg_revenue,
        'avg_engagement': avg_engagement,
        'chart_labels': chart_labels,
        'chart_sales': chart_sales,
        'chart_revenue': chart_revenue,
        'chart_engagement': chart_engagement,
        # Dynamic status messages
        'sales_status': get_performance_status(avg_sales),
        'distribution_status': get_performance_status(avg_distribution),
        'revenue_status': get_performance_status(avg_revenue),
        'engagement_status': get_engagement_status(avg_engagement),
        # Dynamic AI insights
        'team_performance_insight': get_team_performance_insight(avg_sales, avg_distribution, avg_revenue, avg_engagement),
        'growth_trend_insight': get_growth_trend_insight(chart_sales, chart_revenue),
        'recommendation_insight': get_recommendation_insight(avg_engagement, avg_sales, avg_distribution, avg_revenue),
    }

    return render(request, 'dashboard/low_level_manager/analytics.html', context)





def manager_report_dashboard(request):
    records = PerformanceRecord.objects.select_related('employee__user')

    # Filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    department = request.GET.get('department')

    if start_date:
        records = records.filter(performance_start_date__gte=start_date)
    if end_date:
        records = records.filter(performance_end_date__lte=end_date)
    if department:
        records = records.filter(employee__department=department)

    departments = EmployeeProfile.objects.values_list('department', flat=True).distinct()

    # Summary Metrics
    avg_sales = round(sum(r.sales_volume / r.sales_target * 100 for r in records if r.sales_target) / len(records), 2) if records else 0
    avg_revenue = round(sum(r.revenue_volume / r.revenue_target * 100 for r in records if r.revenue_target) / len(records), 2) if records else 0
    avg_engagement = round(sum(r.team_engagement_score for r in records) / len(records), 2) if records else 0

    # Sales Bar Chart Data by Department
    sales_chart = defaultdict(list)
    revenue_chart = defaultdict(float)
    engagement_chart = defaultdict(list)

    for r in records:
        dept = r.employee.department
        if r.sales_target:
            sales_chart[dept].append(r.sales_volume / r.sales_target * 100)
        if r.revenue_target:
            revenue_chart[dept] += r.revenue_volume / r.revenue_target * 100
        engagement_chart[str(r.performance_start_date)].append(r.team_engagement_score)

    sales_chart_data = {
        'labels': list(sales_chart.keys()),
        'data': [round(sum(v)/len(v), 2) for v in sales_chart.values()]
    }

    revenue_chart_data = {
        'labels': list(revenue_chart.keys()),
        'data': [round(v, 2) for v in revenue_chart.values()]
    }

    engagement_chart_data = {
        'labels': list(engagement_chart.keys()),
        'data': [round(sum(v)/len(v), 2) for v in engagement_chart.values()]
    }

    context = {
        'avg_sales': avg_sales,
        'avg_revenue': avg_revenue,
        'avg_engagement': avg_engagement,
        'departments': departments,
        'sales_chart_data': json.dumps(sales_chart_data),
        'revenue_chart_data': json.dumps(revenue_chart_data),
        'engagement_chart_data': json.dumps(engagement_chart_data)
    }

    return render(request, 'dashboard/low_level_manager/reports.html', context)


def export_manager_report(request):
    if request.method == 'POST':
        records = PerformanceRecord.objects.select_related('employee__user')

        # Apply filters
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        department = request.POST.get('department')

        if start_date:
            records = records.filter(performance_start_date__gte=start_date)
        if end_date:
            records = records.filter(performance_end_date__lte=end_date)
        if department:
            records = records.filter(employee__department=department)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="performance_report.csv"'

        writer = csv.writer(response)
        writer.writerow(['Employee ID', 'Department', 'Sales (%)', 'Revenue (%)', 'Engagement', 'Start Date', 'End Date'])

        for r in records:
            sales = round(r.sales_volume / r.sales_target * 100, 2) if r.sales_target else 0
            revenue = round(r.revenue_volume / r.revenue_target * 100, 2) if r.revenue_target else 0
            writer.writerow([
                r.employee.user.employee_id,
                r.employee.department,
                sales,
                revenue,
                r.team_engagement_score,
                r.performance_start_date,
                r.performance_end_date
            ])

        return response





@login_required
def evaluation_list(request):
    evaluations = Evaluation.objects.select_related('employee', 'evaluator').order_by('-date')
    employees = EmployeeProfile.objects.all()
    evaluators = User.objects.filter(is_staff=True)  # assuming managers are staff users

    employee_id = request.GET.get('employee')
    evaluator_id = request.GET.get('evaluator')
    date = request.GET.get('date')

    if employee_id:
        evaluations = evaluations.filter(employee_id=employee_id)
    if evaluator_id:
        evaluations = evaluations.filter(evaluator_id=evaluator_id)
    if date:
        evaluations = evaluations.filter(date=date)

    return render(request, 'dashboard/low_level_manager/evaluation.html', {
        'evaluations': evaluations,
        'employees': employees,
        'evaluators': evaluators,
        'selected_employee': int(employee_id) if employee_id else None,
        'selected_evaluator': int(evaluator_id) if evaluator_id else None,
        'selected_date': date,
    })


@login_required
def add_self_evaluation(request):
    """
    Add self-evaluation for the current user
    """
    try:
        employee_profile = EmployeeProfile.objects.get(user=request.user)
    except EmployeeProfile.DoesNotExist:
        messages.error(request, '❌ Your employee profile is not set up. Please contact HR.')
        return redirect('dashboard')

    if request.method == 'POST':
        try:
            form = EvaluationForm(request.POST)
            if form.is_valid():
                evaluation = form.save(commit=False)
                evaluation.created_by = request.user
                evaluation.evaluator = request.user  # Self-evaluation
                evaluation.evaluation_type = 'SELF'
                evaluation.employee = employee_profile  # Force to be the current user

                # Add evaluation criteria for the employee's role
                employee_role = evaluation.employee.user.role
                criteria = EvaluationCriteria.objects.filter(role=employee_role, is_active=True)
                evaluation.save()  # Save first to get ID

                if criteria.exists():
                    evaluation.evaluation_criteria.set(criteria)

                    # Initialize criteria scores with default values
                    default_scores = {}
                    for criterion in criteria:
                        default_scores[criterion.criteria_name] = 5  # Default score of 5/10
                    evaluation.criteria_scores = default_scores
                    evaluation.save()

                messages.success(request, '✅ Self-evaluation submitted successfully.')
                return redirect('my_evaluation_list')
            else:
                messages.error(request, '❌ Please correct the errors below.')
        except Exception as e:
            messages.error(request, f'❌ Error submitting self-evaluation: {str(e)}')
    else:
        form = EvaluationForm(initial={
            'evaluation_type': 'SELF',
            'employee': employee_profile,
            'evaluator': request.user
        })

    return render(request, 'dashboard/employee/add_self_evaluation.html', {'form': form})


@login_required
def add_manager_evaluation(request, employee_id):
    """
    Add manager evaluation for a specific employee
    """
    if request.user.role not in ['MANAGER', 'MIDDLE_MANAGER', 'HIGH_MANAGER', 'HR']:
        messages.error(request, '❌ You do not have permission to add manager evaluations.')
        return redirect('dashboard')

    try:
        employee = EmployeeProfile.objects.get(id=employee_id)
    except EmployeeProfile.DoesNotExist:
        messages.error(request, '❌ Employee not found.')
        return redirect('employee_management')

    if request.method == 'POST':
        try:
            form = EvaluationForm(request.POST)
            if form.is_valid():
                evaluation = form.save(commit=False)
                evaluation.created_by = request.user
                evaluation.evaluator = request.user
                evaluation.evaluation_type = 'MANAGER'
                evaluation.employee = employee  # Force to be the selected employee

                # Add evaluation criteria for the employee's role
                employee_role = evaluation.employee.user.role
                criteria = EvaluationCriteria.objects.filter(role=employee_role, is_active=True)
                evaluation.save()  # Save first to get ID

                if criteria.exists():
                    evaluation.evaluation_criteria.set(criteria)

                    # Initialize criteria scores with default values
                    default_scores = {}
                    for criterion in criteria:
                        default_scores[criterion.criteria_name] = 5  # Default score of 5/10
                    evaluation.criteria_scores = default_scores
                    evaluation.save()

                messages.success(request, '✅ Manager evaluation submitted successfully.')
                return redirect('evaluation_list')
            else:
                messages.error(request, '❌ Please correct the errors below.')
        except Exception as e:
            messages.error(request, f'❌ Error submitting manager evaluation: {str(e)}')
    else:
        form = EvaluationForm(initial={
            'evaluation_type': 'MANAGER',
            'employee': employee,
            'evaluator': request.user
        })

    return render(request, 'dashboard/low_level_manager/add_manager_evaluation.html', {'form': form, 'employee': employee})


@login_required
def add_360_evaluation(request, employee_id):
    """
    Add 360-degree evaluation combining self, manager, and peer feedback
    """
    if request.user.role not in ['MANAGER', 'MIDDLE_MANAGER', 'HIGH_MANAGER', 'HR']:
        messages.error(request, '❌ You do not have permission to add 360-degree evaluations.')
        return redirect('dashboard')

    try:
        employee = EmployeeProfile.objects.get(id=employee_id)
    except EmployeeProfile.DoesNotExist:
        messages.error(request, '❌ Employee not found.')
        return redirect('employee_management')

    # Get existing evaluations for this employee
    self_evaluation = Evaluation.objects.filter(
        employee=employee,
        evaluation_type='SELF'
    ).order_by('-date').first()

    manager_evaluation = Evaluation.objects.filter(
        employee=employee,
        evaluation_type='MANAGER'
    ).order_by('-date').first()

    peer_reviews = PeerReview.objects.filter(reviewee=employee).order_by('-review_date')[:3]

    if request.method == 'POST':
        try:
            form = EvaluationForm(request.POST)
            if form.is_valid():
                evaluation = form.save(commit=False)
                evaluation.created_by = request.user
                evaluation.evaluator = request.user
                evaluation.evaluation_type = '360'
                evaluation.employee = employee

                # Calculate weighted score from different sources
                weights = {
                    'self': 0.2,      # 20% weight for self-evaluation
                    'manager': 0.5,   # 50% weight for manager evaluation
                    'peer': 0.3       # 30% weight for peer reviews
                }

                weighted_scores = []

                # Self-evaluation score
                if self_evaluation:
                    weighted_scores.append(self_evaluation.performance_score * weights['self'])

                # Manager evaluation score
                if manager_evaluation:
                    weighted_scores.append(manager_evaluation.performance_score * weights['manager'])

                # Peer review average score
                if peer_reviews.exists():
                    peer_avg = sum(review.overall_rating for review in peer_reviews) / peer_reviews.count()
                    weighted_scores.append(peer_avg * weights['peer'])

                # Calculate final weighted score
                if weighted_scores:
                    final_score = sum(weighted_scores) / sum(weights.values()) if weighted_scores else evaluation.performance_score
                    evaluation.performance_score = min(10, max(1, round(final_score)))

                # Add evaluation criteria for the employee's role
                employee_role = evaluation.employee.user.role
                criteria = EvaluationCriteria.objects.filter(role=employee_role, is_active=True)
                evaluation.save()

                if criteria.exists():
                    evaluation.evaluation_criteria.set(criteria)

                    # Initialize criteria scores with weighted averages
                    weighted_criteria_scores = {}
                    for criterion in criteria:
                        scores = []
                        weights_used = []

                        # Self-evaluation criteria score
                        if self_evaluation and self_evaluation.criteria_scores:
                            if criterion.criteria_name in self_evaluation.criteria_scores:
                                scores.append(self_evaluation.criteria_scores[criterion.criteria_name] * weights['self'])
                                weights_used.append(weights['self'])

                        # Manager evaluation criteria score
                        if manager_evaluation and manager_evaluation.criteria_scores:
                            if criterion.criteria_name in manager_evaluation.criteria_scores:
                                scores.append(manager_evaluation.criteria_scores[criterion.criteria_name] * weights['manager'])
                                weights_used.append(weights['manager'])

                        # Peer reviews don't have detailed criteria, so use overall rating
                        if peer_reviews.exists():
                            scores.append((sum(review.overall_rating for review in peer_reviews) / peer_reviews.count()) * weights['peer'])
                            weights_used.append(weights['peer'])

                        if scores:
                            weighted_criteria_scores[criterion.criteria_name] = round(sum(scores) / sum(weights_used))

                    evaluation.criteria_scores = weighted_criteria_scores
                    evaluation.save()

                messages.success(request, '✅ 360-degree evaluation completed successfully.')
                return redirect('evaluation_list')
            else:
                messages.error(request, '❌ Please correct the errors below.')
        except Exception as e:
            messages.error(request, f'❌ Error submitting 360-degree evaluation: {str(e)}')
    else:
        form = EvaluationForm(initial={
            'evaluation_type': '360',
            'employee': employee,
            'evaluator': request.user
        })

    context = {
        'form': form,
        'employee': employee,
        'self_evaluation': self_evaluation,
        'manager_evaluation': manager_evaluation,
        'peer_reviews': peer_reviews,
        'weights': {
            'self': 20,
            'manager': 50,
            'peer': 30
        }
    }

    return render(request, 'dashboard/low_level_manager/add_360_evaluation.html', context)


@login_required
def add_evaluation(request):
    """
    Legacy evaluation view - redirects to appropriate specific evaluation type
    """
    evaluation_type = request.GET.get('type', 'manager')

    if evaluation_type == 'self':
        return redirect('add_self_evaluation')
    elif evaluation_type == '360':
        employee_id = request.GET.get('employee')
        if employee_id:
            return redirect('add_360_evaluation', employee_id=employee_id)
        else:
            messages.error(request, '❌ Employee ID required for 360-degree evaluation.')
    else:
        # Default to manager evaluation
        employee_id = request.GET.get('employee')
        if employee_id:
            return redirect('add_manager_evaluation', employee_id=employee_id)
        else:
            messages.error(request, '❌ Employee ID required for manager evaluation.')

    return redirect('evaluation_list')


@login_required
def edit_evaluation(request, pk):
    evaluation = get_object_or_404(Evaluation, pk=pk)

    # Prevent editing of AI-generated evaluations
    if evaluation.evaluation_type == 'AI':
        messages.error(request, '❌ AI-generated evaluations cannot be edited.')
        return redirect('evaluation_list')

    if request.method == 'POST':
        form = EvaluationForm(request.POST, instance=evaluation)
        if form.is_valid():
            evaluation = form.save(commit=False)
            evaluation.updated_by = request.user
            # Ensure manual evaluations remain marked as such
            evaluation.evaluation_type = 'MANUAL'
            evaluation.save()
            messages.success(request, '✅ Evaluation updated successfully.')
            return redirect('evaluation_list')
        else:
            messages.error(request, '❌ Please correct the errors below.')
    else:
        form = EvaluationForm(instance=evaluation)

    return render(request, 'dashboard/low_level_manager/edit_evaluation.html', {'form': form, 'evaluation': evaluation})


@login_required
def delete_evaluation(request, pk):
    evaluation = get_object_or_404(Evaluation, pk=pk)

    # Prevent deletion of AI-generated evaluations
    if evaluation.evaluation_type == 'AI':
        messages.error(request, '❌ AI-generated evaluations cannot be deleted.')
        return redirect('evaluation_list')

    if request.method == 'POST':
        evaluation.delete()
        messages.success(request, '🗑️ Evaluation deleted successfully.')
        return redirect('evaluation_list')

    messages.error(request, '❌ Invalid request method.')
    return redirect('evaluation_list')





# HIGH LEVEL MANAGER VIEWS
@login_required
@hr_or_admin_required
def deep_analytics(request):
    records = PerformanceRecord.objects.select_related('employee__user')

    # Filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    department = request.GET.get('department')

    if start_date:
        records = records.filter(performance_start_date__gte=start_date)
    if end_date:
        records = records.filter(performance_end_date__lte=end_date)
    if department:
        records = records.filter(employee__department=department)

    departments = EmployeeProfile.objects.values_list('department', flat=True).distinct()

    # Summary Metrics
    avg_sales = round(sum(r.sales_volume / r.sales_target * 100 for r in records if r.sales_target) / len(records), 2) if records else 0
    avg_revenue = round(sum(r.revenue_volume / r.revenue_target * 100 for r in records if r.revenue_target) / len(records), 2) if records else 0
    avg_engagement = round(sum(r.team_engagement_score for r in records) / len(records), 2) if records else 0

    # Charts
    sales_chart = defaultdict(list)
    revenue_chart = defaultdict(float)
    engagement_chart = defaultdict(list)

    for r in records:
        dept = r.employee.department
        if r.sales_target:
            sales_chart[dept].append(r.sales_volume / r.sales_target * 100)
        if r.revenue_target:
            revenue_chart[dept] += r.revenue_volume / r.revenue_target * 100
        engagement_chart[str(r.performance_start_date)].append(r.team_engagement_score)

    sales_chart_data = {
        'labels': list(sales_chart.keys()),
        'data': [round(sum(v)/len(v), 2) for v in sales_chart.values()]
    }

    revenue_chart_data = {
        'labels': list(revenue_chart.keys()),
        'data': [round(v, 2) for v in revenue_chart.values()]
    }

    engagement_chart_data = {
        'labels': list(engagement_chart.keys()),
        'data': [round(sum(v)/len(v), 2) for v in engagement_chart.values()]
    }

    # Evaluation Metrics
    evaluations = Evaluation.objects.select_related('employee__user')
    if department:
        evaluations = evaluations.filter(employee__department=department)

    evaluation_chart = defaultdict(list)
    for e in evaluations:
        full_name = str(e.employee)  # Ensure it's a string
        evaluation_chart[full_name].append(e.performance_score)

    evaluation_chart_data = {
        'labels': list(evaluation_chart.keys()),
        'data': [round(sum(v) / len(v), 2) for v in evaluation_chart.values()]
    }

    avg_evaluation = round(sum(e.performance_score for e in evaluations) / len(evaluations), 2) if evaluations else 0

    context = {
        'avg_sales': avg_sales,
        'avg_revenue': avg_revenue,
        'avg_engagement': avg_engagement,
        'avg_evaluation': avg_evaluation,
        'departments': departments,
        'sales_chart_data': json.dumps(sales_chart_data),
        'revenue_chart_data': json.dumps(revenue_chart_data),
        'engagement_chart_data': json.dumps(engagement_chart_data),
        'evaluation_chart_data': json.dumps(evaluation_chart_data),
    }

    return render(request, 'dashboard/high_level_manager/reports.html', context)





# EMPLOYEE VIEWS
@login_required
def employee(request):
    """
    Enhanced employee dashboard with comprehensive performance analytics
    """
    try:
        employee_profile = EmployeeProfile.objects.get(user=request.user)
    except EmployeeProfile.DoesNotExist:
        messages.error(request, '❌ Your employee profile is not set up. Please contact HR.')
        return redirect('dashboard')

    # Get KPI Scores and Targets
    recent_performance = PerformanceRecord.objects.filter(
        employee=employee_profile
    ).order_by('-performance_end_date').first()

    kpi_data = {}
    if recent_performance:
        # Use new flexible KPI system if available
        role_kpis = KPI.objects.filter(role=employee_profile.user.role, is_active=True)
        if role_kpis.exists():
            for kpi in role_kpis:
                if kpi.name in recent_performance.kpi_values:
                    kpi_info = recent_performance.kpi_values[kpi.name]
                    kpi_data[kpi.name.lower().replace(' ', '_')] = {
                        'name': kpi.name,
                        'current': kpi_info.get('actual', 0),
                        'target': kpi_info.get('target', 0),
                        'percentage': recent_performance.get_kpi_score(kpi.name),
                        'unit': kpi.unit
                    }
        else:
            # Fallback to legacy KPI display
            kpi_data = {
                'sales': {
                    'name': 'Sales',
                    'current': recent_performance.sales_volume or 0,
                    'target': recent_performance.sales_target or 0,
                    'percentage': recent_performance.sales_achieved_percent,
                    'unit': 'USD'
                },
                'distribution': {
                    'name': 'Distribution',
                    'current': recent_performance.distribution_volume or 0,
                    'target': recent_performance.distribution_target or 0,
                    'percentage': recent_performance.distribution_achieved_percent,
                    'unit': 'units'
                },
                'revenue': {
                    'name': 'Revenue',
                    'current': recent_performance.revenue_volume or 0,
                    'target': recent_performance.revenue_target or 0,
                    'percentage': recent_performance.revenue_achieved_percent,
                    'unit': 'USD'
                },
                'customer_base': {
                    'name': 'Customer Base',
                    'current': recent_performance.customer_base_volume or 0,
                    'target': recent_performance.customer_base_target or 0,
                    'percentage': recent_performance.customer_base_achieved_percent,
                    'unit': 'customers'
                },
                'engagement': {
                    'name': 'Team Engagement',
                    'current': recent_performance.team_engagement_score or 0,
                    'target': 8,  # Assuming 8/10 is target
                    'percentage': ((recent_performance.team_engagement_score or 0) / 10) * 100,
                    'unit': '/10'
                }
            }

    # Attendance Analytics for current month
    from django.utils import timezone
    current_month = timezone.now().replace(day=1)
    monthly_attendance = Attendance.objects.filter(
        employee=employee_profile,
        date__gte=current_month
    )

    attendance_data = {
        'present': monthly_attendance.filter(is_present=True).count(),
        'absent': monthly_attendance.filter(is_present=False).count(),
        'late': monthly_attendance.filter(is_late=True).count(),
        'total': monthly_attendance.count()
    }

    # Performance Trend Data (last 6 months)
    from django.db.models.functions import TruncMonth
    from django.db.models import Avg

    # Get performance records for trend
    trend_records = PerformanceRecord.objects.filter(
        employee=employee_profile,
        performance_end_date__gte=timezone.now() - timezone.timedelta(days=180)
    ).order_by('performance_end_date')

    # Calculate trend data
    trend_data = []
    for record in trend_records:
        if hasattr(record, 'get_overall_kpi_score'):
            score = record.get_overall_kpi_score()
        else:
            # Fallback to legacy calculation
            score = (
                (record.sales_achieved_percent / 100) * 0.25 +
                (record.distribution_achieved_percent / 100) * 0.20 +
                (record.revenue_achieved_percent / 100) * 0.30 +
                (record.customer_base_achieved_percent / 100) * 0.15 +
                ((record.team_engagement_score or 0) / 10) * 0.10
            ) * 100

        trend_data.append({
            'month': record.performance_end_date,
            'score': score
        })

    # Group by month
    monthly_trend = {}
    for item in trend_data:
        month_key = item['month'].strftime('%b %Y') if item['month'] else 'Unknown'
        if month_key not in monthly_trend:
            monthly_trend[month_key] = []
        monthly_trend[month_key].append(item['score'])

    performance_trend = {
        'labels': list(monthly_trend.keys()),
        'scores': [round(sum(scores) / len(scores), 1) for scores in monthly_trend.values()]
    }

    # Recent Feedback with Sentiment
    recent_evaluations = Evaluation.objects.filter(
        employee=employee_profile
    ).order_by('-date')[:5]

    feedback_data = []
    for eval in recent_evaluations:
        sentiment = 'neutral'
        if hasattr(eval, 'remarks') and eval.remarks:
            # Simple sentiment analysis based on keywords
            positive_words = ['excellent', 'great', 'good', 'improved', 'strong', 'outstanding']
            negative_words = ['poor', 'weak', 'needs improvement', 'concerning', 'below']

            text = eval.remarks.lower()
            if any(word in text for word in positive_words):
                sentiment = 'positive'
            elif any(word in text for word in negative_words):
                sentiment = 'negative'

        feedback_data.append({
            'evaluator': eval.evaluator.get_full_name() if eval.evaluator else 'Self',
            'date': eval.date,
            'score': eval.performance_score,
            'remarks': eval.remarks[:100] + '...' if eval.remarks and len(eval.remarks) > 100 else eval.remarks,
            'sentiment': sentiment,
            'type': eval.get_evaluation_type_display()
        })

    # Training Progress
    training_data = Training.objects.filter(
        employee=employee_profile
    ).order_by('-start_date')[:5]

    training_progress = []
    for training in training_data:
        training_progress.append({
            'title': training.title,
            'status': training.status,
            'progress': training.completion_percentage,
            'provider': training.provider,
            'skills': training.skills_gained
        })

    # Peer Reviews
    peer_reviews = PeerReview.objects.filter(
        reviewee=employee_profile
    ).select_related('reviewer').order_by('-review_date')[:5]

    # Goals Summary
    tasks_summary = {
        'total': Task.objects.filter(employee=employee_profile).count(),
        'completed': Task.objects.filter(employee=employee_profile, status='COMPLETED').count(),
        'in_progress': Task.objects.filter(employee=employee_profile, status='IN_PROGRESS').count(),
        'overdue': Task.objects.filter(
            employee=employee_profile,
            status__in=['PENDING', 'IN_PROGRESS'],
            due_date__lt=timezone.now().date()
        ).count()
    }

    # Calculate attendance data from database
    current_month = timezone.now().replace(day=1)
    monthly_attendance = Attendance.objects.filter(
        employee=employee_profile,
        date__gte=current_month
    )

    attendance_data = {
        'present': monthly_attendance.filter(is_present=True).count(),
        'absent': monthly_attendance.filter(is_present=False).count(),
        'late': monthly_attendance.filter(is_late=True).count(),
        'total': monthly_attendance.count()
    }

    context = {
        'employee': employee_profile,
        'kpi_data': kpi_data,
        'attendance_data': attendance_data,
        'performance_trend': performance_trend,
        'feedback_data': feedback_data,
        'training_progress': training_progress,
        'tasks_summary': tasks_summary,
        'peer_reviews': peer_reviews,
        'recent_performance': recent_performance
    }

    return render(request, 'dashboard/employee/employee_dashboard.html', context)




@login_required
def my_analytics_view(request):
    user = request.user
    try:
        employee = user.employeeprofile
    except EmployeeProfile.DoesNotExist:
        messages.error(request, '❌ Your employee profile is not set up. Please contact HR.')
        return redirect('dashboard')

    records = PerformanceRecord.objects.filter(employee=employee).order_by('performance_start_date')
    evaluations = Evaluation.objects.filter(employee=employee).order_by('date')

    # Filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        records = records.filter(performance_start_date__gte=start_date)
        evaluations = evaluations.filter(date__gte=start_date)
    if end_date:
        records = records.filter(performance_end_date__lte=end_date)
        evaluations = evaluations.filter(date__lte=end_date)

    # Summary Stats
    avg_sales = round(sum(r.sales_achieved_percent for r in records) / len(records), 2) if records else 0
    avg_revenue = round(sum(r.revenue_achieved_percent for r in records) / len(records), 2) if records else 0
    avg_engagement = round(sum(r.team_engagement_score for r in records) / len(records), 2) if records else 0

    # Chart Data
    performance_chart = {
        'labels': [str(r.performance_start_date) for r in records],
        'sales': [round(r.sales_achieved_percent, 2) for r in records],
        'revenue': [round(r.revenue_achieved_percent, 2) for r in records]
    }

    evaluation_chart = {
        'labels': [str(e.date) for e in evaluations],
        'scores': [e.performance_score for e in evaluations]
    }

    context = {
        'avg_sales': avg_sales,
        'avg_revenue': avg_revenue,
        'avg_engagement': avg_engagement,
        'performance_chart_data': json.dumps(performance_chart),
        'evaluation_chart_data': json.dumps(evaluation_chart),
    }

    return render(request, 'dashboard/employee/emplyee_analytics.html', context)





@login_required
def my_evaluation_list(request):
    evaluations = Evaluation.objects.select_related('employee', 'evaluator').order_by('-date')
    evaluators = User.objects.filter(is_staff=True)  # assuming managers are staff users

    evaluator_id = request.GET.get('evaluator')
    date = request.GET.get('date')

    user = request.user
    try:
        employee = user.employeeprofile
        evaluations = evaluations.filter(employee=employee)
    except EmployeeProfile.DoesNotExist:
        messages.error(request, '❌ Your employee profile is not set up. Please contact HR.')
        return redirect('dashboard')
    if evaluator_id:
        evaluations = evaluations.filter(evaluator_id=evaluator_id)
    if date:
        evaluations = evaluations.filter(date=date)

    return render(request, 'dashboard/employee/my_evaluation.html', {
        'evaluations': evaluations,
        'evaluators': evaluators,
        'selected_evaluator': int(evaluator_id) if evaluator_id else None,
        'selected_date': date,
    })



# AI-POWERED VIEWS

@manager_or_above_required
def ai_employee_analysis(request, employee_id):
    """
    AI-powered comprehensive employee analysis
    """

    try:
        employee = EmployeeProfile.objects.get(id=employee_id)

        # Get recent performance data for simple analysis
        recent_performance = PerformanceRecord.objects.filter(
            employee=employee
        ).order_by('-performance_end_date')[:5]

        # Calculate simple performance metrics
        total_records = recent_performance.count()
        if total_records > 0:
            avg_score = sum(record.get_overall_kpi_score() for record in recent_performance if hasattr(record, 'get_overall_kpi_score')) / total_records
        else:
            avg_score = 0

        # Get recent evaluations
        recent_evaluations = Evaluation.objects.filter(
            employee=employee
        ).order_by('-date')[:3]

        context = {
            'employee': employee,
            'recent_performance': recent_performance,
            'avg_performance_score': round(avg_score, 1),
            'recent_evaluations': recent_evaluations,
            'total_records': total_records,
        }

        return render(request, 'dashboard/ai/employee_analysis.html', context)

    except EmployeeProfile.DoesNotExist:
        messages.error(request, '❌ Employee not found.')
        return redirect('employee_management')
    except Exception as e:
        messages.error(request, f'❌ Analysis error: {str(e)}')
        return redirect('employee_management')

@hr_or_admin_required
def ai_department_analysis(request):
    """
    AI-powered department analysis
    """

    department = request.GET.get('department')
    if not department:
        departments = EmployeeProfile.objects.values_list('department', flat=True).distinct()
        return render(request, 'dashboard/ai/department_select.html', {'departments': departments})

    try:
        ai_service = AIService()
        ai_service.load_active_models()
        analysis_result = ai_service.analyze_department_performance(department, user=request.user)

        if 'error' in analysis_result:
            messages.error(request, f'❌ Department analysis failed: {analysis_result["error"]}')
            return redirect('ai_department_analysis')

        context = {
            'department': department,
            'analysis': analysis_result,
            'total_employees': analysis_result.get('total_employees', 0),
            'performance_distribution': analysis_result.get('performance_distribution', {}),
            'top_performers': analysis_result.get('top_performers', []),
            'needs_attention': analysis_result.get('needs_attention', []),
            'department_insights': analysis_result.get('department_insights', []),
        }

        return render(request, 'dashboard/ai/department_analysis.html', context)

    except Exception as e:
        messages.error(request, f'❌ Analysis error: {str(e)}')
        return redirect('ai_department_analysis')

@login_required
def ai_recommendations_dashboard(request):
    """
    Dashboard showing AI-generated recommendations
    """
    from ai_engine.models import Recommendation

    if request.user.role not in ['HR', 'HIGH_MANAGER', 'MIDDLE_MANAGER', 'MANAGER']:
        messages.error(request, '❌ You do not have permission to access AI recommendations.')
        return redirect('dashboard')

    # Get recommendations based on user role
    recommendations = Recommendation.objects.select_related('employee').order_by('-created_at')

    if request.user.role == 'MANAGER':
        # Managers see recommendations for their direct reports
        recommendations = recommendations.filter(employee__user__role='EMPLOYEE')
    elif request.user.role == 'MIDDLE_MANAGER':
        recommendations = recommendations.filter(employee__user__role__in=['EMPLOYEE', 'MANAGER'])
    # HR and HIGH_MANAGER can see all

    # Filter by priority if specified
    priority = request.GET.get('priority')
    if priority:
        recommendations = recommendations.filter(priority=priority)

    # Filter by category if specified
    category = request.GET.get('category')
    if category:
        recommendations = recommendations.filter(category=category)

    context = {
        'recommendations': recommendations,
        'selected_priority': priority,
        'selected_category': category,
        'priorities': ['LOW', 'MEDIUM', 'HIGH', 'CRITICAL'],
        'categories': ['sales', 'engagement', 'development', 'performance'],
    }

    return render(request, 'dashboard/ai/recommendations_dashboard.html', context)

@login_required
def ai_insights_dashboard(request):
    """
    Dashboard showing AI-generated insights
    """
    from ai_engine.models import PerformanceInsight

    if request.user.role not in ['HR', 'HIGH_MANAGER']:
        messages.error(request, '❌ You do not have permission to access AI insights.')
        return redirect('dashboard')

    insights = PerformanceInsight.objects.select_related('generated_by').order_by('-created_at')

    # Filter by type if specified
    insight_type = request.GET.get('type')
    if insight_type:
        insights = insights.filter(insight_type=insight_type)

    context = {
        'insights': insights,
        'selected_type': insight_type,
        'insight_types': ['TREND', 'COMPARISON', 'FORECAST', 'CLUSTER'],
    }

    return render(request, 'dashboard/ai/insights_dashboard.html', context)

@login_required
def ai_anomalies_dashboard(request):
    """
    Dashboard showing detected performance anomalies
    """
    from ai_engine.models import Anomaly

    if request.user.role not in ['HR', 'HIGH_MANAGER', 'MIDDLE_MANAGER']:
        messages.error(request, '❌ You do not have permission to access anomaly detection.')
        return redirect('dashboard')

    anomalies = Anomaly.objects.select_related('employee').order_by('-detected_at')

    # Filter by acknowledged status
    acknowledged = request.GET.get('acknowledged')
    if acknowledged == 'true':
        anomalies = anomalies.filter(is_acknowledged=True)
    elif acknowledged == 'false':
        anomalies = anomalies.filter(is_acknowledged=False)

    context = {
        'anomalies': anomalies,
        'selected_acknowledged': acknowledged,
    }

    return render(request, 'dashboard/ai/anomalies_dashboard.html', context)


@login_required
@hr_or_admin_required
def hr_dashboard(request):
    """
    HR Executive Dashboard with real-time organization-wide analytics
    """
    from django.db.models import Avg, Count, Q, F, ExpressionWrapper, FloatField
    from django.db.models.functions import TruncMonth

    # Basic organization statistics
    total_employees = EmployeeProfile.objects.count()
    departments_count = EmployeeProfile.objects.values('department').distinct().count()

    # Department performance data - calculate in Python since sales_achieved_percent is a property
    all_records = PerformanceRecord.objects.select_related('employee').filter(employee__department__isnull=False)
    dept_data = {}

    for record in all_records:
        dept = record.employee.department
        if dept not in dept_data:
            dept_data[dept] = {'scores': [], 'employee_count': 0}
        dept_data[dept]['scores'].append(record.sales_achieved_percent)
        dept_data[dept]['employee_count'] = len(set(r.employee.id for r in all_records if r.employee.department == dept))

    department_performance = []
    for dept, data in dept_data.items():
        if data['scores']:
            avg_performance = sum(data['scores']) / len(data['scores'])
            department_performance.append({
                'employee__department': dept,
                'avg_performance': avg_performance,
                'employee_count': data['employee_count']
            })

    # Sort by average performance descending
    department_performance.sort(key=lambda x: x['avg_performance'], reverse=True)

    dept_labels = [item['employee__department'] for item in department_performance]
    dept_scores = [round(item['avg_performance'] or 0, 1) for item in department_performance]

    # Performance distribution calculation
    all_performance_records = PerformanceRecord.objects.all()
    performance_scores = []

    for record in all_performance_records:
        if hasattr(record, 'get_overall_kpi_score'):
            score = record.get_overall_kpi_score()
        else:
            # Fallback to legacy calculation
            score = record.sales_achieved_percent
        performance_scores.append(score)

    # Calculate distribution
    if performance_scores:
        high_performers = sum(1 for score in performance_scores if score >= 90)
        average_performers = sum(1 for score in performance_scores if 70 <= score < 90)
        needs_attention = sum(1 for score in performance_scores if 50 <= score < 70)
        top_talent = sum(1 for score in performance_scores if score >= 95)

        total_records = len(performance_scores)
        perf_distribution = {
            'high_performers': round((high_performers / total_records) * 100),
            'average': round((average_performers / total_records) * 100),
            'needs_attention': round((needs_attention / total_records) * 100),
            'top_talent': round((top_talent / total_records) * 100),
        }
    else:
        perf_distribution = {
            'high_performers': 0,
            'average': 0,
            'needs_attention': 0,
            'top_talent': 0,
        }

    # Organization health metrics
    # Employee satisfaction (based on evaluation scores)
    avg_evaluation_score = Evaluation.objects.aggregate(
        avg_score=Avg('performance_score')
    )['avg_score'] or 0

    # Retention rate (employees with recent activity)
    from django.utils import timezone
    six_months_ago = timezone.now() - timezone.timedelta(days=180)
    active_employees = EmployeeProfile.objects.filter(
        Q(user__last_login__gte=six_months_ago) |
        Q(data_updated__gte=six_months_ago) |
        Q(performance_records__performance_end_date__gte=six_months_ago)
    ).distinct().count()

    retention_rate = round((active_employees / total_employees) * 100) if total_employees > 0 else 0

    # Engagement score (based on team engagement scores)
    avg_engagement = PerformanceRecord.objects.aggregate(
        avg_engagement=Avg('team_engagement_score')
    )['avg_engagement'] or 0

    # Diversity index (simplified - based on department distribution)
    dept_distribution = EmployeeProfile.objects.values('department').annotate(
        count=Count('id')
    ).filter(department__isnull=False)

    if dept_distribution:
        # Calculate diversity as 1 - (max_dept / total_employees)^2
        max_dept_size = max(item['count'] for item in dept_distribution)
        diversity_index = round((1 - (max_dept_size / total_employees) ** 2) * 100)
    else:
        diversity_index = 0

    # Top performers
    top_performers = []
    if performance_scores:
        # Get employees with highest performance scores by sorting in Python
        all_records = PerformanceRecord.objects.select_related('employee')
        sorted_records = sorted(all_records, key=lambda r: r.sales_achieved_percent, reverse=True)[:3]

        for record in sorted_records:
            top_performers.append({
                'name': f"{record.employee.first_name} {record.employee.last_name}",
                'score': round(record.sales_achieved_percent, 1),
                'department': record.employee.job_title or 'Employee'
            })

    # Development opportunities
    # Based on training needs and low performance areas
    trainings_needed = Training.objects.filter(status__in=['PLANNED', 'IN_PROGRESS']).count()
    low_performers = sum(1 for score in performance_scores if score < 70)

    development_needs = [
        {'type': 'Leadership Training', 'count': max(1, round(trainings_needed * 0.3))},
        {'type': 'Technical Skills', 'count': max(1, round(trainings_needed * 0.4))},
        {'type': 'Soft Skills Development', 'count': max(1, round(trainings_needed * 0.3))},
    ]

    # Organization trend data (last 12 months)
    from django.db.models.functions import ExtractMonth, ExtractYear

    # Calculate monthly performance in Python since sales_achieved_percent is a property
    monthly_records = PerformanceRecord.objects.annotate(
        month=ExtractMonth('performance_end_date'),
        year=ExtractYear('performance_end_date')
    ).values('year', 'month', 'sales_volume', 'sales_target')

    monthly_data = {}
    for record in monthly_records:
        key = f"{record['year']}-{record['month']:02d}"
        if key not in monthly_data:
            monthly_data[key] = []

        # Calculate performance percentage
        if record['sales_target'] and record['sales_target'] > 0:
            performance = (record['sales_volume'] / record['sales_target']) * 100
        else:
            performance = 0
        monthly_data[key].append(performance)

    # Calculate averages
    monthly_performance = []
    for key, performances in monthly_data.items():
        if performances:
            year, month = key.split('-')
            monthly_performance.append({
                'year': int(year),
                'month': int(month),
                'avg_performance': sum(performances) / len(performances)
            })

    # Sort by year and month
    monthly_performance.sort(key=lambda x: (x['year'], x['month']))

    # Create trend data for the last 12 months
    trend_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    trend_scores = []

    # Get current year data
    current_year = timezone.now().year
    for i, month in enumerate(trend_labels, 1):
        # Find data for this month in our calculated monthly_performance
        month_data = None
        for data in monthly_performance:
            if data['year'] == current_year and data['month'] == i:
                month_data = data
                break

        if month_data:
            trend_scores.append(round(month_data['avg_performance'] or 0, 1))
        else:
            # Use average of available data or 0
            trend_scores.append(round(sum(performance_scores) / len(performance_scores), 1) if performance_scores else 0)

    context = {
        'total_employees': total_employees,
        'departments_count': departments_count,
        'department_labels': dept_labels,
        'department_scores': dept_scores,
        'performance_distribution': perf_distribution,
        'organization_health': {
            'satisfaction': round(avg_evaluation_score * 10),  # Convert to percentage
            'retention': retention_rate,
            'engagement': round(avg_engagement, 1),
            'diversity': diversity_index,
        },
        'top_performers': top_performers,
        'development_needs': development_needs,
        'trend_labels': trend_labels,
        'trend_scores': trend_scores,
    }

    return render(request, 'dashboard/hr/hr_dashboard.html', context)


@login_required
def ai_demo_dashboard(request):
    """
    AI Features Demonstration Dashboard - Perfect for project defense
    Showcases all AI capabilities in one comprehensive view
    """
    if request.user.role not in ['HR', 'HIGH_MANAGER']:
        messages.error(request, '❌ You do not have permission to access AI demonstration.')
        return redirect('dashboard')

    from ai_engine.inference_service import AIService
    from performance_app.models import EmployeeProfile

    ai_service = AIService()

    # Get a sample employee for demonstration (first employee with performance data)
    sample_employee = EmployeeProfile.objects.filter(
        user__role='EMPLOYEE'
    ).first()

    demo_data = {}

    if sample_employee:
        # Run comprehensive AI analysis
        analysis = ai_service.analyze_employee_performance(sample_employee.id, user=request.user)

        if 'error' not in analysis:
            demo_data = {
                'employee': sample_employee,
                'analysis': analysis,
                'ai_features': {
                    'sentiment_analysis': {
                        'title': 'Sentiment Analysis with Hugging Face',
                        'description': 'Uses pre-trained RoBERTa model to analyze feedback sentiment',
                        'demo_data': analysis.get('sentiment_analysis', {}),
                        'icon': 'fas fa-heart',
                        'color': 'pink'
                    },
                    'ai_summary': {
                        'title': 'AI-Generated Performance Summary',
                        'description': 'Natural language generation of performance insights',
                        'demo_data': analysis.get('ai_performance_summary', ''),
                        'icon': 'fas fa-brain',
                        'color': 'indigo'
                    },
                    'promotion_assessment': {
                        'title': 'Promotion Readiness Assessment',
                        'description': 'Rule-based AI assessment of career advancement potential',
                        'demo_data': analysis.get('promotion_readiness', {}),
                        'icon': 'fas fa-trophy',
                        'color': 'green'
                    },
                    'smart_recommendations': {
                        'title': 'Smart Recommendation Engine',
                        'description': 'AI-powered suggestions for training and development',
                        'demo_data': analysis.get('ai_recommendations', []),
                        'icon': 'fas fa-lightbulb',
                        'color': 'yellow'
                    },
                    'performance_prediction': {
                        'title': 'Performance Prediction',
                        'description': 'Machine learning-based future performance forecasting',
                        'demo_data': analysis.get('predictions', []),
                        'icon': 'fas fa-chart-line',
                        'color': 'blue'
                    }
                }
            }

    context = {
        'demo_data': demo_data,
        'ai_technologies': [
            {
                'name': 'Hugging Face Transformers',
                'purpose': 'Sentiment Analysis & Text Processing',
                'model': 'RoBERTa-base for sentiment classification'
            },
            {
                'name': 'TextBlob',
                'purpose': 'Rule-based sentiment analysis fallback',
                'model': 'Lexicon-based approach'
            },
            {
                'name': 'Scikit-learn',
                'purpose': 'Machine Learning for predictions',
                'model': 'Regression models for performance forecasting'
            },
            {
                'name': 'Natural Language Generation',
                'purpose': 'AI text summary generation',
                'model': 'Template-based with dynamic content'
            }
        ]
    }

    return render(request, 'dashboard/ai/ai_demo_dashboard.html', context)


# ATTENDANCE MANAGEMENT VIEWS
@login_required
def attendance_list(request):
    """
    List attendance records with filtering
    """
    user = request.user
    records = Attendance.objects.select_related('employee').order_by('-date')

    # Role-based filtering
    if user.role == 'EMPLOYEE':
        records = records.filter(employee__user=user)
    elif user.role in ['MANAGER', 'MIDDLE_MANAGER']:
        # Managers see their team's attendance
        team_members = EmployeeProfile.objects.filter(
            Q(user__role='EMPLOYEE') |
            Q(user__role='MANAGER') |
            Q(user__role='MIDDLE_MANAGER')
        )
        records = records.filter(employee__in=team_members)

    # Filters
    employee_id = request.GET.get('employee')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status = request.GET.get('status')

    if employee_id:
        records = records.filter(employee_id=employee_id)
    if start_date:
        records = records.filter(date__gte=start_date)
    if end_date:
        records = records.filter(date__lte=end_date)
    if status:
        if status == 'present':
            records = records.filter(is_present=True)
        elif status == 'absent':
            records = records.filter(is_present=False)
        elif status == 'late':
            records = records.filter(is_late=True)
        elif status == 'early_departure':
            records = records.filter(is_early_departure=True)

    employees = EmployeeProfile.objects.all() if user.role in ['HR', 'HIGH_MANAGER'] else None

    context = {
        'records': records,
        'employees': employees,
        'selected_employee': employee_id,
        'start_date': start_date,
        'end_date': end_date,
        'selected_status': status,
    }

    return render(request, 'dashboard/attendance/attendance_list.html', context)


@login_required
def add_attendance(request):
    """
    Add attendance record
    """
    if request.user.role not in ['HR', 'HIGH_MANAGER', 'MANAGER', 'MIDDLE_MANAGER']:
        messages.error(request, '❌ You do not have permission to add attendance records.')
        return redirect('dashboard')

    if request.method == 'POST':
        employee_id = request.POST.get('employee')
        date = request.POST.get('date')
        check_in = request.POST.get('check_in_time')
        check_out = request.POST.get('check_out_time')
        is_present = request.POST.get('is_present') == 'on'
        notes = request.POST.get('notes', '')

        try:
            employee = EmployeeProfile.objects.get(id=employee_id)

            # Check if record already exists
            existing = Attendance.objects.filter(employee=employee, date=date).first()
            if existing:
                messages.error(request, '❌ Attendance record already exists for this date.')
                return redirect('attendance_list')

            attendance = Attendance.objects.create(
                employee=employee,
                date=date,
                check_in_time=check_in if check_in else None,
                check_out_time=check_out if check_out else None,
                is_present=is_present,
                notes=notes,
                created_by=request.user
            )

            # Calculate derived fields
            if attendance.check_in_time and attendance.check_out_time:
                # Simple calculation - can be enhanced
                hours = (attendance.check_out_time.hour - attendance.check_in_time.hour) + \
                       (attendance.check_out_time.minute - attendance.check_in_time.minute) / 60.0
                attendance.hours_worked = max(0, hours)
                attendance.save()

            messages.success(request, '✅ Attendance record added successfully.')
            return redirect('attendance_list')

        except EmployeeProfile.DoesNotExist:
            messages.error(request, '❌ Employee not found.')
        except Exception as e:
            messages.error(request, f'❌ Error adding attendance: {str(e)}')

    employees = EmployeeProfile.objects.all()
    return render(request, 'dashboard/attendance/add_attendance.html', {'employees': employees})


# TASK MANAGEMENT VIEWS
@login_required
def task_list(request):
    """
    List tasks with filtering
    """
    user = request.user
    tasks = Task.objects.select_related('employee').order_by('-assigned_date')

    # Role-based filtering
    if user.role == 'EMPLOYEE':
        tasks = tasks.filter(employee__user=user)
    elif user.role in ['MANAGER', 'MIDDLE_MANAGER']:
        team_members = EmployeeProfile.objects.filter(
            Q(user__role='EMPLOYEE') |
            Q(user__role='MANAGER')
        )
        tasks = tasks.filter(employee__in=team_members)

    # Filters
    employee_id = request.GET.get('employee')
    status = request.GET.get('status')
    priority = request.GET.get('priority')

    if employee_id:
        tasks = tasks.filter(employee_id=employee_id)
    if status:
        tasks = tasks.filter(status=status)
    if priority:
        tasks = tasks.filter(priority=priority)

    employees = EmployeeProfile.objects.all() if user.role in ['HR', 'HIGH_MANAGER'] else None

    context = {
        'tasks': tasks,
        'employees': employees,
        'selected_employee': employee_id,
        'selected_status': status,
        'selected_priority': priority,
    }

    return render(request, 'dashboard/tasks/task_list.html', context)


@login_required
def add_task(request):
    """
    Add new task
    """
    if request.user.role not in ['HR', 'HIGH_MANAGER', 'MANAGER', 'MIDDLE_MANAGER']:
        messages.error(request, '❌ You do not have permission to add tasks.')
        return redirect('dashboard')

    if request.method == 'POST':
        try:
            employee = EmployeeProfile.objects.get(id=request.POST.get('employee'))
            task = Task.objects.create(
                employee=employee,
                title=request.POST.get('title'),
                description=request.POST.get('description'),
                priority=request.POST.get('priority'),
                due_date=request.POST.get('due_date') or None,
                estimated_hours=request.POST.get('estimated_hours') or None,
                project_name=request.POST.get('project_name'),
                created_by=request.user
            )
            messages.success(request, '✅ Task added successfully.')
            return redirect('task_list')
        except Exception as e:
            messages.error(request, f'❌ Error adding task: {str(e)}')

    employees = EmployeeProfile.objects.all()
    return render(request, 'dashboard/tasks/add_task.html', {'employees': employees})


@login_required
def update_task_status(request, task_id):
    """
    Update task status (for employees and managers)
    """
    if request.method == 'POST':
        try:
            task = Task.objects.get(id=task_id)
            user = request.user

            # Check permissions
            if user.role == 'EMPLOYEE' and task.employee.user != user:
                messages.error(request, '❌ You can only update your own tasks.')
                return redirect('task_list')

            new_status = request.POST.get('status')
            progress = request.POST.get('progress_percentage')
            actual_hours = request.POST.get('actual_hours')

            task.status = new_status
            if progress:
                task.progress_percentage = int(progress)
            if actual_hours:
                task.actual_hours = float(actual_hours)
            if new_status == 'COMPLETED':
                task.completed_date = timezone.now().date()

            task.updated_by = user
            task.save()

            messages.success(request, '✅ Task updated successfully.')
        except Task.DoesNotExist:
            messages.error(request, '❌ Task not found.')
        except Exception as e:
            messages.error(request, f'❌ Error updating task: {str(e)}')

    return redirect('task_list')


# PEER REVIEW VIEWS
@login_required
def peer_review_list(request):
    """
    List peer reviews
    """
    user = request.user
    reviews = PeerReview.objects.select_related('reviewer', 'reviewee').order_by('-review_date')

    # Role-based filtering
    if user.role == 'EMPLOYEE':
        reviews = reviews.filter(Q(reviewer__user=user) | Q(reviewee__user=user))

    context = {'reviews': reviews}
    return render(request, 'peer_reviews/peer_review_list.html', context)


@login_required
def add_peer_review(request):
    """
    Add peer review
    """
    if request.method == 'POST':
        try:
            # Get reviewer profile with error handling
            try:
                reviewer = EmployeeProfile.objects.get(user=request.user)
            except EmployeeProfile.DoesNotExist:
                messages.error(request, '❌ Your employee profile is not set up. Please contact HR.')
                return redirect('peer_review_list')

            # Get reviewee profile with error handling
            reviewee_id = request.POST.get('reviewee')
            if not reviewee_id:
                messages.error(request, '❌ Please select an employee to review.')
                return redirect('peer_review_list')

            try:
                reviewee = EmployeeProfile.objects.get(id=reviewee_id)
            except EmployeeProfile.DoesNotExist:
                messages.error(request, '❌ Selected employee profile not found.')
                return redirect('peer_review_list')

            if reviewer == reviewee:
                messages.error(request, '❌ You cannot review yourself.')
                return redirect('peer_review_list')

            review = PeerReview.objects.create(
                reviewer=reviewer,
                reviewee=reviewee,
                strengths=request.POST.get('strengths'),
                areas_for_improvement=request.POST.get('areas_for_improvement'),
                overall_rating=int(request.POST.get('overall_rating')),
                collaboration_score=int(request.POST.get('collaboration_score')),
                communication_score=int(request.POST.get('communication_score')),
                technical_skills_score=int(request.POST.get('technical_skills_score')),
                feedback=request.POST.get('feedback'),
                is_anonymous=request.POST.get('is_anonymous') == 'on',
                created_by=request.user
            )

            messages.success(request, '✅ Peer review submitted successfully.')
            return redirect('peer_review_list')

        except Exception as e:
            messages.error(request, f'❌ Error submitting review: {str(e)}')

    # Get potential reviewees (exclude self and people user has already reviewed recently)
    try:
        reviewer_profile = EmployeeProfile.objects.get(user=request.user)
        recent_reviews = PeerReview.objects.filter(
            reviewer=reviewer_profile,
            review_date__gte=timezone.now().date() - timedelta(days=30)
        ).values_list('reviewee_id', flat=True)

        employees = EmployeeProfile.objects.exclude(
            Q(user=request.user) | Q(id__in=recent_reviews)
        )
    except EmployeeProfile.DoesNotExist:
        # If reviewer profile doesn't exist, show empty list
        employees = EmployeeProfile.objects.none()
        messages.warning(request, '⚠️ Your employee profile is not set up. Please contact HR to submit peer reviews.')

    return render(request, 'peer_reviews/add_peer_review.html', {'employees': employees})


# TRAINING MANAGEMENT VIEWS
@login_required
def training_list(request):
    """
    List training records
    """
    user = request.user
    trainings = Training.objects.select_related('employee').order_by('-start_date')

    # Role-based filtering
    if user.role == 'EMPLOYEE':
        trainings = trainings.filter(employee__user=user)

    # Filters
    employee_id = request.GET.get('employee')
    status = request.GET.get('status')
    training_type = request.GET.get('training_type')

    if employee_id:
        trainings = trainings.filter(employee_id=employee_id)
    if status:
        trainings = trainings.filter(status=status)
    if training_type:
        trainings = trainings.filter(training_type__icontains=training_type)

    employees = EmployeeProfile.objects.all() if user.role in ['HR', 'HIGH_MANAGER'] else None

    context = {
        'trainings': trainings,
        'employees': employees,
        'selected_employee': employee_id,
        'selected_status': status,
        'selected_type': training_type,
    }

    return render(request, 'training/training_list.html', context)


@login_required
def add_training(request):
    """
    Add training record
    """
    if request.user.role not in ['HR', 'HIGH_MANAGER', 'MANAGER', 'MIDDLE_MANAGER']:
        messages.error(request, '❌ You do not have permission to add training records.')
        return redirect('dashboard')

    if request.method == 'POST':
        try:
            employee = EmployeeProfile.objects.get(id=request.POST.get('employee'))
            training = Training.objects.create(
                employee=employee,
                title=request.POST.get('title'),
                description=request.POST.get('description'),
                training_type=request.POST.get('training_type'),
                provider=request.POST.get('provider'),
                start_date=request.POST.get('start_date') or None,
                end_date=request.POST.get('end_date') or None,
                cost=request.POST.get('cost') or None,
                notes=request.POST.get('notes'),
                created_by=request.user
            )
            messages.success(request, '✅ Training record added successfully.')
            return redirect('training_list')
        except Exception as e:
            messages.error(request, f'❌ Error adding training: {str(e)}')

    employees = EmployeeProfile.objects.all()
    return render(request, 'training/add_training.html', {'employees': employees})


@login_required
def update_training_progress(request, training_id):
    """
    Update training progress
    """
    if request.method == 'POST':
        try:
            training = Training.objects.get(id=training_id)
            user = request.user

            # Check permissions
            if user.role == 'EMPLOYEE' and training.employee.user != user:
                messages.error(request, '❌ You can only update your own training.')
                return redirect('training_list')

            progress = request.POST.get('completion_percentage')
            status = request.POST.get('status')
            skills = request.POST.get('skills_gained')
            certification = request.POST.get('certification_earned')

            if progress:
                training.completion_percentage = int(progress)
            if status:
                training.status = status
                if status == 'COMPLETED':
                    training.end_date = timezone.now().date()
            if skills:
                training.skills_gained = skills.split(',')
            if certification:
                training.certification_earned = certification

            training.updated_by = user
            training.save()

            messages.success(request, '✅ Training updated successfully.')
        except Training.DoesNotExist:
            messages.error(request, '❌ Training not found.')
        except Exception as e:
            messages.error(request, f'❌ Error updating training: {str(e)}')

    return redirect('training_list')


# EVALUATION CRITERIA MANAGEMENT
@login_required
def kpi_list(request):
    """
    List KPIs by role
    """
    if request.user.role not in ['HR', 'HIGH_MANAGER']:
        messages.error(request, '❌ You do not have permission to manage KPIs.')
        return redirect('dashboard')

    role = request.GET.get('role')
    kpis = KPI.objects.all()

    if role:
        kpis = kpis.filter(role=role)

    roles = User.ROLE_CHOICES
    context = {
        'kpis': kpis,
        'roles': roles,
        'selected_role': role,
    }

    return render(request, 'dashboard/kpi/kpi_list.html', context)


@login_required
def add_kpi(request):
    """
    Add new KPI
    """
    if request.user.role not in ['HR', 'HIGH_MANAGER']:
        messages.error(request, '❌ You do not have permission to add KPIs.')
        return redirect('dashboard')

    if request.method == 'POST':
        try:
            kpi = KPI.objects.create(
                role=request.POST.get('role'),
                name=request.POST.get('name'),
                description=request.POST.get('description'),
                kpi_type=request.POST.get('kpi_type'),
                unit=request.POST.get('unit'),
                target_direction=request.POST.get('target_direction'),
                weight=request.POST.get('weight'),
                created_by=request.user
            )
            messages.success(request, '✅ KPI added successfully.')
            return redirect('kpi_list')
        except Exception as e:
            messages.error(request, f'❌ Error adding KPI: {str(e)}')

    roles = User.ROLE_CHOICES
    return render(request, 'dashboard/kpi/add_kpi.html', {'roles': roles})


@login_required
def edit_kpi(request, kpi_id):
    """
    Edit existing KPI
    """
    if request.user.role not in ['HR', 'HIGH_MANAGER']:
        messages.error(request, '❌ You do not have permission to edit KPIs.')
        return redirect('dashboard')

    kpi = get_object_or_404(KPI, id=kpi_id)

    if request.method == 'POST':
        try:
            kpi.role = request.POST.get('role')
            kpi.name = request.POST.get('name')
            kpi.description = request.POST.get('description')
            kpi.kpi_type = request.POST.get('kpi_type')
            kpi.unit = request.POST.get('unit')
            kpi.target_direction = request.POST.get('target_direction')
            kpi.weight = request.POST.get('weight')
            kpi.updated_by = request.user
            kpi.save()
            messages.success(request, '✅ KPI updated successfully.')
            return redirect('kpi_list')
        except Exception as e:
            messages.error(request, f'❌ Error updating KPI: {str(e)}')

    roles = User.ROLE_CHOICES
    return render(request, 'dashboard/kpi/edit_kpi.html', {'kpi': kpi, 'roles': roles})


@login_required
def delete_kpi(request, kpi_id):
    """
    Delete KPI
    """
    if request.user.role not in ['HR', 'HIGH_MANAGER']:
        messages.error(request, '❌ You do not have permission to delete KPIs.')
        return redirect('dashboard')

    kpi = get_object_or_404(KPI, id=kpi_id)

    if request.method == 'POST':
        kpi.delete()
        messages.success(request, '🗑️ KPI deleted successfully.')
        return redirect('kpi_list')

    return render(request, 'dashboard/kpi/delete_kpi.html', {'kpi': kpi})


@login_required
def evaluation_criteria_list(request):
    """
    List evaluation criteria by role
    """
    if request.user.role not in ['HR', 'HIGH_MANAGER']:
        messages.error(request, '❌ You do not have permission to manage evaluation criteria.')
        return redirect('dashboard')

    role = request.GET.get('role')
    criteria = EvaluationCriteria.objects.all()

    if role:
        criteria = criteria.filter(role=role)

    roles = User.ROLE_CHOICES
    context = {
        'criteria': criteria,
        'roles': roles,
        'selected_role': role,
    }

    return render(request, 'dashboard/evaluation_criteria/criteria_list.html', context)


@login_required
def add_evaluation_criteria(request):
    """
    Add evaluation criteria
    """
    if request.user.role not in ['HR', 'HIGH_MANAGER']:
        messages.error(request, '❌ You do not have permission to add evaluation criteria.')
        return redirect('dashboard')

    if request.method == 'POST':
        try:
            criteria = EvaluationCriteria.objects.create(
                role=request.POST.get('role'),
                criteria_name=request.POST.get('criteria_name'),
                description=request.POST.get('description'),
                weight=request.POST.get('weight'),
                created_by=request.user
            )
            messages.success(request, '✅ Evaluation criteria added successfully.')
            return redirect('evaluation_criteria_list')
        except Exception as e:
            messages.error(request, f'❌ Error adding criteria: {str(e)}')

    roles = User.ROLE_CHOICES
    return render(request, 'dashboard/evaluation_criteria/add_criteria.html', {'roles': roles})


# ENHANCED ANALYTICS AND CHART VIEWS
@login_required
def attendance_analytics(request):
    """
    Comprehensive attendance analytics with charts
    """
    user = request.user

    # Base queryset
    attendance_records = Attendance.objects.select_related('employee')

    # Role-based filtering
    if user.role == 'EMPLOYEE':
        attendance_records = attendance_records.filter(employee__user=user)
    elif user.role in ['MANAGER', 'MIDDLE_MANAGER']:
        # Managers see their team's attendance
        team_members = EmployeeProfile.objects.filter(
            Q(user__role='EMPLOYEE') |
            Q(user__role='MANAGER')
        )
        attendance_records = attendance_records.filter(employee__in=team_members)

    # Filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    employee_id = request.GET.get('employee')

    if start_date:
        attendance_records = attendance_records.filter(date__gte=start_date)
    if end_date:
        attendance_records = attendance_records.filter(date__lte=end_date)
    if employee_id:
        attendance_records = attendance_records.filter(employee_id=employee_id)

    # Calculate attendance metrics
    total_records = attendance_records.count()
    present_count = attendance_records.filter(is_present=True).count()
    absent_count = attendance_records.filter(is_present=False).count()
    late_count = attendance_records.filter(is_late=True).count()
    early_departure_count = attendance_records.filter(is_early_departure=True).count()

    attendance_rate = (present_count / total_records * 100) if total_records > 0 else 0
    punctuality_rate = ((total_records - late_count) / total_records * 100) if total_records > 0 else 0

    # Daily attendance trend (last 30 days)
    from django.db.models.functions import TruncDate
    from django.db.models import Count

    daily_trend = attendance_records.filter(
        date__gte=timezone.now().date() - timedelta(days=30)
    ).annotate(
        date_trunc=TruncDate('date')
    ).values('date_trunc').annotate(
        total=Count('id'),
        present=Count('id', filter=Q(is_present=True)),
        late=Count('id', filter=Q(is_late=True))
    ).order_by('date_trunc')

    # Prepare chart data
    trend_labels = []
    trend_present = []
    trend_total = []
    trend_late = []

    for day in daily_trend:
        trend_labels.append(day['date_trunc'].strftime('%Y-%m-%d'))
        trend_present.append(day['present'])
        trend_total.append(day['total'])
        trend_late.append(day['late'])

    # Department-wise attendance
    dept_attendance = attendance_records.values('employee__department').annotate(
        total=Count('id'),
        present=Count('id', filter=Q(is_present=True)),
        attendance_rate=ExpressionWrapper(
            Count('id', filter=Q(is_present=True)) * 100.0 / Count('id'),
            output_field=FloatField()
        )
    ).order_by('-attendance_rate')

    dept_labels = [item['employee__department'] for item in dept_attendance]
    dept_rates = [round(item['attendance_rate'], 1) for item in dept_attendance]

    # Employee attendance ranking
    employee_ranking = attendance_records.values(
        'employee__first_name', 'employee__last_name', 'employee__user__employee_id'
    ).annotate(
        total_days=Count('id'),
        present_days=Count('id', filter=Q(is_present=True)),
        attendance_rate=ExpressionWrapper(
            Count('id', filter=Q(is_present=True)) * 100.0 / Count('id'),
            output_field=FloatField()
        )
    ).order_by('-attendance_rate')[:10]

    employee_labels = [f"{emp['employee__first_name']} {emp['employee__last_name']}" for emp in employee_ranking]
    employee_rates = [round(emp['attendance_rate'], 1) for emp in employee_ranking]

    employees = EmployeeProfile.objects.all() if user.role in ['HR', 'HIGH_MANAGER'] else None

    context = {
        'total_records': total_records,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'early_departure_count': early_departure_count,
        'attendance_rate': round(attendance_rate, 1),
        'punctuality_rate': round(punctuality_rate, 1),
        'trend_labels': json.dumps(trend_labels),
        'trend_present': json.dumps(trend_present),
        'trend_total': json.dumps(trend_total),
        'trend_late': json.dumps(trend_late),
        'dept_labels': json.dumps(dept_labels),
        'dept_rates': json.dumps(dept_rates),
        'employee_labels': json.dumps(employee_labels),
        'employee_rates': json.dumps(employee_rates),
        'employees': employees,
        'selected_employee': employee_id,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'dashboard/analytics/attendance_analytics.html', context)


@login_required
def task_analytics(request):
    """
    Task completion analytics and charts
    """
    user = request.user

    # Base queryset
    tasks = Task.objects.select_related('employee')

    # Role-based filtering
    if user.role == 'EMPLOYEE':
        tasks = tasks.filter(employee__user=user)
    elif user.role in ['MANAGER', 'MIDDLE_MANAGER']:
        team_members = EmployeeProfile.objects.filter(
            Q(user__role='EMPLOYEE') |
            Q(user__role='MANAGER')
        )
        tasks = tasks.filter(employee__in=team_members)

    # Task status distribution
    status_distribution = tasks.values('status').annotate(
        count=Count('id')
    ).order_by('status')

    status_labels = [item['status'] for item in status_distribution]
    status_counts = [item['count'] for item in status_distribution]

    # Priority distribution
    priority_distribution = tasks.values('priority').annotate(
        count=Count('id')
    ).order_by('priority')

    priority_labels = [item['priority'] for item in priority_distribution]
    priority_counts = [item['count'] for item in priority_distribution]

    # Completion rate by employee
    employee_completion = tasks.values(
        'employee__first_name', 'employee__last_name'
    ).annotate(
        total_tasks=Count('id'),
        completed_tasks=Count('id', filter=Q(status='COMPLETED')),
        completion_rate=ExpressionWrapper(
            Count('id', filter=Q(status='COMPLETED')) * 100.0 / Count('id'),
            output_field=FloatField()
        )
    ).order_by('-completion_rate')[:10]

    emp_labels = [f"{emp['employee__first_name']} {emp['employee__last_name']}" for emp in employee_completion]
    emp_rates = [round(emp['completion_rate'], 1) for emp in employee_completion]

    # Monthly task completion trend
    monthly_completion = tasks.filter(
        assigned_date__gte=timezone.now().date() - timedelta(days=180)
    ).extra(
        select={'month': "DATE_FORMAT(assigned_date, '%%Y-%%m')"}
    ).values('month').annotate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='COMPLETED'))
    ).order_by('month')

    month_labels = [item['month'] for item in monthly_completion]
    month_total = [item['total'] for item in monthly_completion]
    month_completed = [item['completed'] for item in monthly_completion]

    # Overdue tasks analysis
    overdue_tasks = tasks.filter(
        status__in=['PENDING', 'IN_PROGRESS'],
        due_date__lt=timezone.now().date()
    )

    overdue_by_priority = overdue_tasks.values('priority').annotate(
        count=Count('id')
    ).order_by('priority')

    overdue_labels = [item['priority'] for item in overdue_by_priority]
    overdue_counts = [item['count'] for item in overdue_by_priority]

    context = {
        'total_tasks': tasks.count(),
        'completed_tasks': tasks.filter(status='COMPLETED').count(),
        'in_progress_tasks': tasks.filter(status='IN_PROGRESS').count(),
        'overdue_tasks': overdue_tasks.count(),
        'status_labels': json.dumps(status_labels),
        'status_counts': json.dumps(status_counts),
        'priority_labels': json.dumps(priority_labels),
        'priority_counts': json.dumps(priority_counts),
        'emp_labels': json.dumps(emp_labels),
        'emp_rates': json.dumps(emp_rates),
        'month_labels': json.dumps(month_labels),
        'month_total': json.dumps(month_total),
        'month_completed': json.dumps(month_completed),
        'overdue_labels': json.dumps(overdue_labels),
        'overdue_counts': json.dumps(overdue_counts),
    }

    return render(request, 'dashboard/analytics/task_analytics.html', context)


@login_required
def training_analytics(request):
    """
    Training completion and effectiveness analytics
    """
    user = request.user

    # Base queryset
    trainings = Training.objects.select_related('employee')

    # Role-based filtering
    if user.role == 'EMPLOYEE':
        trainings = trainings.filter(employee__user=user)

    # Training status distribution
    status_distribution = trainings.values('status').annotate(
        count=Count('id')
    ).order_by('status')

    status_labels = [item['status'] for item in status_distribution]
    status_counts = [item['count'] for item in status_distribution]

    # Training type distribution
    type_distribution = trainings.values('training_type').annotate(
        count=Count('id')
    ).order_by('-count')[:10]

    type_labels = [item['training_type'] for item in type_distribution]
    type_counts = [item['count'] for item in type_distribution]

    # Completion rate by training type
    completion_by_type = trainings.values('training_type').annotate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='COMPLETED')),
        completion_rate=ExpressionWrapper(
            Count('id', filter=Q(status='COMPLETED')) * 100.0 / Count('id'),
            output_field=FloatField()
        )
    ).order_by('-completion_rate')

    comp_type_labels = [item['training_type'] for item in completion_by_type]
    comp_type_rates = [round(item['completion_rate'], 1) for item in completion_by_type]

    # Monthly training completion
    monthly_training = trainings.filter(
        start_date__gte=timezone.now().date() - timedelta(days=180)
    ).extra(
        select={'month': "DATE_FORMAT(start_date, '%%Y-%%m')"}
    ).values('month').annotate(
        started=Count('id'),
        completed=Count('id', filter=Q(status='COMPLETED'))
    ).order_by('month')

    train_month_labels = [item['month'] for item in monthly_training]
    train_started = [item['started'] for item in monthly_training]
    train_completed = [item['completed'] for item in monthly_training]

    # Cost analysis
    total_cost = trainings.aggregate(
        total=Sum('cost', filter=Q(cost__isnull=False))
    )['total'] or 0

    avg_cost_per_training = trainings.filter(cost__isnull=False).aggregate(
        avg=Avg('cost')
    )['avg'] or 0

    context = {
        'total_trainings': trainings.count(),
        'completed_trainings': trainings.filter(status='COMPLETED').count(),
        'in_progress_trainings': trainings.filter(status='IN_PROGRESS').count(),
        'total_cost': round(total_cost, 2),
        'avg_cost_per_training': round(avg_cost_per_training, 2),
        'status_labels': json.dumps(status_labels),
        'status_counts': json.dumps(status_counts),
        'type_labels': json.dumps(type_labels),
        'type_counts': json.dumps(type_counts),
        'comp_type_labels': json.dumps(comp_type_labels),
        'comp_type_rates': json.dumps(comp_type_rates),
        'train_month_labels': json.dumps(train_month_labels),
        'train_started': json.dumps(train_started),
        'train_completed': json.dumps(train_completed),
    }

    return render(request, 'dashboard/analytics/training_analytics.html', context)


# COMPREHENSIVE REPORTING SYSTEM
@login_required
def generate_performance_report(request):
    """
    Generate comprehensive performance reports in PDF/Excel/CSV formats
    """
    user = request.user
    report_type = request.GET.get('type', 'pdf')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    employee_id = request.GET.get('employee')
    department = request.GET.get('department')

    # Base querysets
    performance_records = PerformanceRecord.objects.select_related('employee__user')
    evaluations = Evaluation.objects.select_related('employee__user', 'evaluator')
    attendance_records = Attendance.objects.select_related('employee__user')
    tasks = Task.objects.select_related('employee__user')
    trainings = Training.objects.select_related('employee__user')

    # Apply filters
    if start_date:
        performance_records = performance_records.filter(performance_start_date__gte=start_date)
        evaluations = evaluations.filter(date__gte=start_date)
        attendance_records = attendance_records.filter(date__gte=start_date)
        tasks = tasks.filter(assigned_date__gte=start_date)
        trainings = trainings.filter(start_date__gte=start_date)

    if end_date:
        performance_records = performance_records.filter(performance_end_date__lte=end_date)
        evaluations = evaluations.filter(date__lte=end_date)
        attendance_records = attendance_records.filter(date__lte=end_date)
        tasks = tasks.filter(due_date__lte=end_date)
        trainings = trainings.filter(end_date__lte=end_date)

    if employee_id:
        performance_records = performance_records.filter(employee_id=employee_id)
        evaluations = evaluations.filter(employee_id=employee_id)
        attendance_records = attendance_records.filter(employee_id=employee_id)
        tasks = tasks.filter(employee_id=employee_id)
        trainings = trainings.filter(employee_id=employee_id)

    if department:
        performance_records = performance_records.filter(employee__department=department)
        evaluations = evaluations.filter(employee__department=department)
        attendance_records = attendance_records.filter(employee__department=department)
        tasks = tasks.filter(employee__department=department)
        trainings = trainings.filter(employee__department=department)

    # Role-based access control
    if user.role == 'EMPLOYEE':
        performance_records = performance_records.filter(employee__user=user)
        evaluations = evaluations.filter(employee__user=user)
        attendance_records = attendance_records.filter(employee__user=user)
        tasks = tasks.filter(employee__user=user)
        trainings = trainings.filter(employee__user=user)
    elif user.role in ['MANAGER', 'MIDDLE_MANAGER']:
        # Managers can see their team's data
        team_members = EmployeeProfile.objects.filter(
            Q(user__role='EMPLOYEE') |
            Q(user__role='MANAGER')
        )
        performance_records = performance_records.filter(employee__in=team_members)
        evaluations = evaluations.filter(employee__in=team_members)
        attendance_records = attendance_records.filter(employee__in=team_members)
        tasks = tasks.filter(employee__in=team_members)
        trainings = trainings.filter(employee__in=team_members)

    if report_type == 'pdf':
        return generate_pdf_report(
            performance_records, evaluations, attendance_records, tasks, trainings,
            start_date, end_date, department
        )
    elif report_type == 'excel':
        return generate_excel_report(
            performance_records, evaluations, attendance_records, tasks, trainings,
            start_date, end_date, department
        )
    else:  # CSV
        return generate_csv_report(
            performance_records, evaluations, attendance_records, tasks, trainings,
            start_date, end_date, department
        )


def generate_pdf_report(performance_records, evaluations, attendance_records, tasks, trainings,
                       start_date, end_date, department):
    """
    Generate PDF report using ReportLab
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
    )
    title = Paragraph("Employee Performance Report", title_style)
    elements.append(title)

    # Report period
    period_text = f"Report Period: {start_date or 'All'} to {end_date or 'Present'}"
    if department:
        period_text += f" | Department: {department}"
    elements.append(Paragraph(period_text, styles['Normal']))
    elements.append(Spacer(1, 12))

    # Performance Summary
    elements.append(Paragraph("Performance Summary", styles['Heading2']))

    perf_data = [['Employee', 'Avg Sales %', 'Avg Revenue %', 'Avg Engagement', 'Evaluations']]
    for record in performance_records.values('employee__first_name', 'employee__last_name').distinct():
        emp_name = f"{record['employee__first_name']} {record['employee__last_name']}"
        emp_records = performance_records.filter(
            employee__first_name=record['employee__first_name'],
            employee__last_name=record['employee__last_name']
        )

        # Calculate averages in Python since these are properties, not database fields
        sales_percentages = [r.sales_achieved_percent for r in emp_records]
        revenue_percentages = [r.revenue_achieved_percent for r in emp_records]
        engagement_scores = [r.team_engagement_score or 0 for r in emp_records]

        avg_sales = sum(sales_percentages) / len(sales_percentages) if sales_percentages else 0
        avg_revenue = sum(revenue_percentages) / len(revenue_percentages) if revenue_percentages else 0
        avg_engagement = sum(engagement_scores) / len(engagement_scores) if engagement_scores else 0

        eval_count = evaluations.filter(
            employee__first_name=record['employee__first_name'],
            employee__last_name=record['employee__last_name']
        ).count()

        perf_data.append([
            emp_name,
            f"{avg_sales:.1f}%",
            f"{avg_revenue:.1f}%",
            f"{avg_engagement:.1f}",
            str(eval_count)
        ])

    perf_table = Table(perf_data)
    perf_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(perf_table)
    elements.append(Spacer(1, 20))

    # Attendance Summary
    elements.append(Paragraph("Attendance Summary", styles['Heading2']))

    attendance_data = [['Employee', 'Total Days', 'Present', 'Absent', 'Late', 'Attendance %']]
    for record in attendance_records.values('employee__first_name', 'employee__last_name').distinct():
        emp_name = f"{record['employee__first_name']} {record['employee__last_name']}"
        emp_attendance = attendance_records.filter(
            employee__first_name=record['employee__first_name'],
            employee__last_name=record['employee__last_name']
        )
        total = emp_attendance.count()
        present = emp_attendance.filter(is_present=True).count()
        absent = emp_attendance.filter(is_present=False).count()
        late = emp_attendance.filter(is_late=True).count()
        attendance_rate = (present / total * 100) if total > 0 else 0

        attendance_data.append([
            emp_name, str(total), str(present), str(absent), str(late), f"{attendance_rate:.1f}%"
        ])

    attendance_table = Table(attendance_data)
    attendance_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(attendance_table)

    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="performance_report.pdf"'
    return response


def generate_excel_report(performance_records, evaluations, attendance_records, tasks, trainings,
                         start_date, end_date, department):
    """
    Generate Excel report using openpyxl
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Report"

    # Title
    ws['A1'] = "Employee Performance Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Report Period: {start_date or 'All'} to {end_date or 'Present'}"
    if department:
        ws['A3'] = f"Department: {department}"

    # Performance Summary
    ws['A5'] = "Performance Summary"
    ws['A5'].font = Font(size=14, bold=True)

    headers = ['Employee', 'Avg Sales %', 'Avg Revenue %', 'Avg Engagement', 'Evaluations']
    for col, header in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=header).font = Font(bold=True)
        ws.cell(row=6, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    row = 7
    for record in performance_records.values('employee__first_name', 'employee__last_name').distinct():
        emp_name = f"{record['employee__first_name']} {record['employee__last_name']}"
        emp_records = performance_records.filter(
            employee__first_name=record['employee__first_name'],
            employee__last_name=record['employee__last_name']
        )

        # Calculate averages in Python since these are properties, not database fields
        sales_percentages = [r.sales_achieved_percent for r in emp_records]
        revenue_percentages = [r.revenue_achieved_percent for r in emp_records]
        engagement_scores = [r.team_engagement_score or 0 for r in emp_records]

        avg_sales = sum(sales_percentages) / len(sales_percentages) if sales_percentages else 0
        avg_revenue = sum(revenue_percentages) / len(revenue_percentages) if revenue_percentages else 0
        avg_engagement = sum(engagement_scores) / len(engagement_scores) if engagement_scores else 0

        eval_count = evaluations.filter(
            employee__first_name=record['employee__first_name'],
            employee__last_name=record['employee__last_name']
        ).count()

        ws.cell(row=row, column=1, value=emp_name)
        ws.cell(row=row, column=2, value=round(avg_sales, 1))
        ws.cell(row=row, column=3, value=round(avg_revenue, 1))
        ws.cell(row=row, column=4, value=round(avg_engagement, 1))
        ws.cell(row=row, column=5, value=eval_count)
        row += 1

    # Attendance Summary
    row += 2
    ws.cell(row=row, column=1, value="Attendance Summary").font = Font(size=14, bold=True)
    row += 1

    attendance_headers = ['Employee', 'Total Days', 'Present', 'Absent', 'Late', 'Attendance %']
    for col, header in enumerate(attendance_headers, 1):
        ws.cell(row=row, column=col, value=header).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    row += 1
    for record in attendance_records.values('employee__first_name', 'employee__last_name').distinct():
        emp_name = f"{record['employee__first_name']} {record['employee__last_name']}"
        emp_attendance = attendance_records.filter(
            employee__first_name=record['employee__first_name'],
            employee__last_name=record['employee__last_name']
        )
        total = emp_attendance.count()
        present = emp_attendance.filter(is_present=True).count()
        absent = emp_attendance.filter(is_present=False).count()
        late = emp_attendance.filter(is_late=True).count()
        attendance_rate = (present / total * 100) if total > 0 else 0

        ws.cell(row=row, column=1, value=emp_name)
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=3, value=present)
        ws.cell(row=row, column=4, value=absent)
        ws.cell(row=row, column=5, value=late)
        ws.cell(row=row, column=6, value=round(attendance_rate, 1))
        row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="performance_report.xlsx"'
    return response


def generate_csv_report(performance_records, evaluations, attendance_records, tasks, trainings,
                        start_date, end_date, department):
    """
    Generate CSV report (enhanced version of existing)
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="performance_report.csv"'

    writer = csv.writer(response)
    writer.writerow(['Employee Performance Report'])
    writer.writerow([f'Report Period: {start_date or "All"} to {end_date or "Present"}'])
    if department:
        writer.writerow([f'Department: {department}'])
    writer.writerow([])

    # Performance Summary
    writer.writerow(['PERFORMANCE SUMMARY'])
    writer.writerow(['Employee', 'Avg Sales %', 'Avg Revenue %', 'Avg Engagement', 'Evaluations'])

    for record in performance_records.values('employee__first_name', 'employee__last_name').distinct():
        emp_name = f"{record['employee__first_name']} {record['employee__last_name']}"
        emp_records = performance_records.filter(
            employee__first_name=record['employee__first_name'],
            employee__last_name=record['employee__last_name']
        )

        # Calculate averages in Python since these are properties, not database fields
        sales_percentages = [r.sales_achieved_percent for r in emp_records]
        revenue_percentages = [r.revenue_achieved_percent for r in emp_records]
        engagement_scores = [r.team_engagement_score or 0 for r in emp_records]

        avg_sales = sum(sales_percentages) / len(sales_percentages) if sales_percentages else 0
        avg_revenue = sum(revenue_percentages) / len(revenue_percentages) if revenue_percentages else 0
        avg_engagement = sum(engagement_scores) / len(engagement_scores) if engagement_scores else 0

        eval_count = evaluations.filter(
            employee__first_name=record['employee__first_name'],
            employee__last_name=record['employee__last_name']
        ).count()

        writer.writerow([
            emp_name,
            f"{avg_sales:.1f}%",
            f"{avg_revenue:.1f}%",
            f"{avg_engagement:.1f}",
            str(eval_count)
        ])

    writer.writerow([])

    # Attendance Summary
    writer.writerow(['ATTENDANCE SUMMARY'])
    writer.writerow(['Employee', 'Total Days', 'Present', 'Absent', 'Late', 'Attendance %'])

    for record in attendance_records.values('employee__first_name', 'employee__last_name').distinct():
        emp_name = f"{record['employee__first_name']} {record['employee__last_name']}"
        emp_attendance = attendance_records.filter(
            employee__first_name=record['employee__first_name'],
            employee__last_name=record['employee__last_name']
        )
        total = emp_attendance.count()
        present = emp_attendance.filter(is_present=True).count()
        absent = emp_attendance.filter(is_present=False).count()
        late = emp_attendance.filter(is_late=True).count()
        attendance_rate = (present / total * 100) if total > 0 else 0

        writer.writerow([
            emp_name, str(total), str(present), str(absent), str(late), f"{attendance_rate:.1f}%"
        ])

    return response


@login_required
@hr_or_admin_required
def export_organization_report(request):
    """
    Enhanced organization-wide export functionality
    """
    report_type = request.GET.get('type', 'excel')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    department = request.GET.get('department')

    # Get all data for organization report
    performance_records = PerformanceRecord.objects.select_related('employee__user')
    evaluations = Evaluation.objects.select_related('employee__user', 'evaluator')
    attendance_records = Attendance.objects.select_related('employee__user')
    tasks = Task.objects.select_related('employee__user')
    trainings = Training.objects.select_related('employee__user')

    # Apply filters
    if start_date:
        performance_records = performance_records.filter(performance_start_date__gte=start_date)
        evaluations = evaluations.filter(date__gte=start_date)
        attendance_records = attendance_records.filter(date__gte=start_date)
        tasks = tasks.filter(assigned_date__gte=start_date)
        trainings = trainings.filter(start_date__gte=start_date)

    if end_date:
        performance_records = performance_records.filter(performance_end_date__lte=end_date)
        evaluations = evaluations.filter(date__lte=end_date)
        attendance_records = attendance_records.filter(date__lte=end_date)
        tasks = tasks.filter(due_date__lte=end_date)
        trainings = trainings.filter(end_date__lte=end_date)

    if department:
        performance_records = performance_records.filter(employee__department=department)
        evaluations = evaluations.filter(employee__department=department)
        attendance_records = attendance_records.filter(employee__department=department)
        tasks = tasks.filter(employee__department=department)
        trainings = trainings.filter(employee__department=department)

    if report_type == 'pdf':
        return generate_organization_pdf_report(
            performance_records, evaluations, attendance_records, tasks, trainings,
            start_date, end_date, department
        )
    elif report_type == 'excel':
        return generate_organization_excel_report(
            performance_records, evaluations, attendance_records, tasks, trainings,
            start_date, end_date, department
        )
    else:  # CSV
        return generate_organization_csv_report(
            performance_records, evaluations, attendance_records, tasks, trainings,
            start_date, end_date, department
        )


def generate_organization_pdf_report(performance_records, evaluations, attendance_records, tasks, trainings,
                                   start_date, end_date, department):
    """
    Generate comprehensive organization PDF report
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
    )
    title = Paragraph("Organization Performance Report", title_style)
    elements.append(title)

    # Report period
    period_text = f"Report Period: {start_date or 'All'} to {end_date or 'Present'}"
    if department:
        period_text += f" | Department: {department}"
    elements.append(Paragraph(period_text, styles['Normal']))
    elements.append(Spacer(1, 12))

    # Organization Summary
    elements.append(Paragraph("Organization Summary", styles['Heading2']))

    # Calculate summary stats
    total_employees = performance_records.values('employee').distinct().count()

    # Calculate average performance in Python since it's a property
    performance_scores = [r.sales_achieved_percent for r in performance_records]
    avg_performance = sum(performance_scores) / len(performance_scores) if performance_scores else 0

    total_evaluations = evaluations.count()
    avg_attendance = 0
    if attendance_records.exists():
        total_days = attendance_records.count()
        present_days = attendance_records.filter(is_present=True).count()
        avg_attendance = (present_days / total_days * 100) if total_days > 0 else 0

    summary_data = [
        ['Metric', 'Value'],
        ['Total Employees', str(total_employees)],
        ['Average Performance', f"{avg_performance:.1f}%"],
        ['Total Evaluations', str(total_evaluations)],
        ['Average Attendance', f"{avg_attendance:.1f}%"],
    ]

    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Top Performers
    elements.append(Paragraph("Top Performers", styles['Heading2']))

    top_performers_data = [['Employee', 'Performance %', 'Department']]

    # Get top performers by calculating performance in Python
    all_records = list(performance_records)
    if all_records:
        # Sort by sales_achieved_percent in descending order
        sorted_records = sorted(all_records, key=lambda r: r.sales_achieved_percent, reverse=True)[:5]

        for record in sorted_records:
            top_performers_data.append([
                f"{record.employee.first_name} {record.employee.last_name}",
                f"{record.sales_achieved_percent:.1f}%",
                record.employee.department
            ])

    top_table = Table(top_performers_data)
    top_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(top_table)

    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="organization_report.pdf"'
    return response


def generate_organization_excel_report(performance_records, evaluations, attendance_records, tasks, trainings,
                                     start_date, end_date, department):
    """
    Generate comprehensive organization Excel report
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Organization Report"

    # Title
    ws['A1'] = "Organization Performance Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Report Period: {start_date or 'All'} to {end_date or 'Present'}"
    if department:
        ws['A3'] = f"Department: {department}"

    # Summary Section
    ws['A5'] = "Organization Summary"
    ws['A5'].font = Font(size=14, bold=True)

    summary_headers = ['Metric', 'Value']
    for col, header in enumerate(summary_headers, 1):
        ws.cell(row=6, column=col, value=header).font = Font(bold=True)
        ws.cell(row=6, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Calculate summary stats
    total_employees = performance_records.values('employee').distinct().count()

    # Calculate average performance in Python since it's a property
    performance_scores = [r.sales_achieved_percent for r in performance_records]
    avg_performance = sum(performance_scores) / len(performance_scores) if performance_scores else 0

    total_evaluations = evaluations.count()

    summary_data = [
        ["Total Employees", total_employees],
        ["Average Performance", f"{avg_performance:.1f}%"],
        ["Total Evaluations", total_evaluations],
    ]

    for row, data in enumerate(summary_data, 7):
        ws.cell(row=row, column=1, value=data[0])
        ws.cell(row=row, column=2, value=data[1])

    # Performance Details Section
    ws['A12'] = "Performance Details"
    ws['A12'].font = Font(size=14, bold=True)

    detail_headers = ['Employee', 'Department', 'Performance %', 'Evaluations', 'Attendance %']
    for col, header in enumerate(detail_headers, 1):
        ws.cell(row=13, column=col, value=header).font = Font(bold=True)
        ws.cell(row=13, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    row = 14
    for record in performance_records:
        emp_evaluations = evaluations.filter(employee=record.employee).count()
        emp_attendance = attendance_records.filter(employee=record.employee)
        attendance_rate = 0
        if emp_attendance.exists():
            total_days = emp_attendance.count()
            present_days = emp_attendance.filter(is_present=True).count()
            attendance_rate = (present_days / total_days * 100) if total_days > 0 else 0

        ws.cell(row=row, column=1, value=f"{record.employee.first_name} {record.employee.last_name}")
        ws.cell(row=row, column=2, value=record.employee.department)
        ws.cell(row=row, column=3, value=f"{record.sales_achieved_percent:.1f}%")
        ws.cell(row=row, column=4, value=emp_evaluations)
        ws.cell(row=row, column=5, value=f"{attendance_rate:.1f}%")
        row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="organization_report.xlsx"'
    return response


def generate_organization_csv_report(performance_records, evaluations, attendance_records, tasks, trainings,
                                   start_date, end_date, department):
    """
    Generate comprehensive organization CSV report
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="organization_report.csv"'

    writer = csv.writer(response)
    writer.writerow(['Organization Performance Report'])
    writer.writerow([f'Report Period: {start_date or "All"} to {end_date or "Present"}'])
    if department:
        writer.writerow([f'Department: {department}'])
    writer.writerow([])

    # Organization Summary
    writer.writerow(['ORGANIZATION SUMMARY'])
    total_employees = performance_records.values('employee').distinct().count()

    # Calculate average performance in Python since it's a property
    performance_scores = [r.sales_achieved_percent for r in performance_records]
    avg_performance = sum(performance_scores) / len(performance_scores) if performance_scores else 0

    total_evaluations = evaluations.count()

    writer.writerow(['Total Employees', total_employees])
    writer.writerow(['Average Performance', f"{avg_performance:.1f}%"])
    writer.writerow(['Total Evaluations', total_evaluations])
    writer.writerow([])

    # Detailed Performance Data
    writer.writerow(['EMPLOYEE PERFORMANCE DETAILS'])
    writer.writerow(['Employee', 'Department', 'Performance %', 'Evaluations', 'Attendance %'])

    for record in performance_records:
        emp_evaluations = evaluations.filter(employee=record.employee).count()
        emp_attendance = attendance_records.filter(employee=record.employee)
        attendance_rate = 0
        if emp_attendance.exists():
            total_days = emp_attendance.count()
            present_days = emp_attendance.filter(is_present=True).count()
            attendance_rate = (present_days / total_days * 100) if total_days > 0 else 0

        writer.writerow([
            f"{record.employee.first_name} {record.employee.last_name}",
            record.employee.department,
            f"{record.sales_achieved_percent:.1f}%",
            emp_evaluations,
            f"{attendance_rate:.1f}%"
        ])

    return response